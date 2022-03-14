from openpyxl import load_workbook, Workbook
from openpyxl.styles.numbers import BUILTIN_FORMATS
from models import HoleNorm, Part, PointPart, SawNorm, AssemblyNorm, WeldNorm, Point, Drawing, Order, User, Worker
from peewee import fn


def AAA():
  wb = load_workbook('users.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('users')
  post = []
  for y in range(1,14):
    d = []
    for i in range(1,5):
      d.append(sheet.cell(row=y,column=i).value)
    post.append(d)
  User.insert_many(post, fields=[User.id,User.surname,User.name,User.patronymic]).execute()

def BBB():
  wb = load_workbook('workers.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('workers')
  post = []
  for y in range(1,15):
    d = []
    for i in range(1,5):
      d.append(sheet.cell(row=y,column=i).value)
    post.append(d)
  Worker.insert_many(post, fields=[Worker.id,Worker.user,Worker.oper,Worker.oper_rus]).execute()

def CCC():
  for i in PointPart.select(PointPart,Part,Point).join(Part).join(Drawing).join(Order).join_from(PointPart,Point).where(Order.cas == '2325',Part.number == 280,Point.faza == 8):
    print(i.detail)
  