from openpyxl import load_workbook, Workbook
from openpyxl.styles.numbers import BUILTIN_FORMATS
from models import HoleNorm, Part, PointPart, SawNorm, AssemblyNorm, WeldNorm, Point, Drawing, Order, User, Worker
from peewee import fn

def NormExcel():
  HoleNormExcel()
  SawNormExcel()
  AssemblyNormExcel()
  WeldNormExcel()
  # User_worker()

def HoleNormExcel():
  wb = load_workbook('VVL.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Сверление')
  post = []
  for y in range(2,164):
    d = []
    for i in range(1,7):
      if i == 1:
        d.append((sheet.cell(row=y,column=i).value).split('-')[0])
        d.append((sheet.cell(row=y,column=i).value).split('-')[1])
      elif i == 2 and sheet.cell(row=y,column=i).value == 33:
        d.append(1000)
      elif i == 3 and sheet.cell(row=y,column=i).value == 'до 3':
        d.append(0)
        d.append(2999)
      elif i == 3 and sheet.cell(row=y,column=i).value == 'свыше 3':
        d.append(3000)
        d.append(30000)
      elif i == 3 and sheet.cell(row=y,column=i).value == 0:
        d.append(0)
        d.append(0)
      else:
        d.append(sheet.cell(row=y,column=i).value)
    post.append(d)
  HoleNorm.insert_many(post, fields=[HoleNorm.depth_of,HoleNorm.depth_to,HoleNorm.diameter,HoleNorm.lenght_of,HoleNorm.lenght_to,HoleNorm.count,HoleNorm.norm,HoleNorm.metal]).execute()


def SawNormExcel():
  wb = load_workbook('VVL.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Пиление проката')
  post = []
  for y in range(2,161):
    d = []
    for i in range(1,7):
      if i == 1 and len((sheet.cell(row=y,column=i).value).strip().split(' ')) > 2:
        d.append((sheet.cell(row=y,column=i).value).strip().split(' ')[0] + ' ' + (sheet.cell(row=y,column=i).value).split(' ')[1])
        d.append((sheet.cell(row=y,column=i).value).strip().split(' ')[2])
      elif i == 1 and len((sheet.cell(row=y,column=i).value).strip().split(' ')) == 2:
        d.append((sheet.cell(row=y,column=i).value).strip().split(' ')[0])
        d.append((sheet.cell(row=y,column=i).value).strip().split(' ')[1])
      elif sheet.cell(row=y,column=i).value == None:
        d.append(0)
      else:
        d.append(sheet.cell(row=y,column=i).value)
    post.append(d)
  SawNorm.insert_many(post, fields=[SawNorm.profile,SawNorm.size,SawNorm.speed_saw,SawNorm.speed_feed,SawNorm.step_tooth,SawNorm.norm_direct,SawNorm.norm_oblique]).execute()


def AssemblyNormExcel():
  wb = load_workbook('VVL.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Сборка (ЕНиР)')
  post = []
  for y in range(17,865):
    d = []
    for i in range(1,9):
      if sheet.cell(row=y,column=i).value == None:
        d.append('')
      else:
        d.append(sheet.cell(row=y,column=i).value)
    post.append(d)
  AssemblyNorm.insert_many(post, fields=[AssemblyNorm.name,AssemblyNorm.mass_of,AssemblyNorm.mass_to,AssemblyNorm.count_of,AssemblyNorm.count_to,AssemblyNorm.complexity,AssemblyNorm.norm,AssemblyNorm.choice]).execute()


def WeldNormExcel():
  wb = load_workbook('VVL.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Сварка')
  post = []
  for y in range(1,12):
    d = []
    for i in range(1,3):
      if sheet.cell(row=y,column=i).value == None:
        d.append('')
      else:
        d.append(sheet.cell(row=y,column=i).value)
    post.append(d)
  WeldNorm.insert_many(post, fields=[WeldNorm.cathet,WeldNorm.norm]).execute()


def Statement(faza,case):
  book = Workbook()
  sheet = book.active
  stat = PointPart.select(PointPart,Drawing,Point,(fn.COUNT(Part.id) / fn.COUNT(fn.DISTINCT(Part.id))).alias('count')).join(Point).join(Drawing).join(Order).join_from(PointPart,Part).where(Order.cas == case,Point.faza == faza).group_by(Point.assembly,PointPart.detail)
  y = 1
  sheet.cell(row=y,column=1).value = 'Наряд'
  sheet.cell(row=y,column=2).value = 'Марка'
  sheet.cell(row=y,column=3).value = 'Наименование'
  sheet.cell(row=y,column=4).value = 'Чертеж'
  sheet.cell(row=y,column=5).value = 'Количество'
  sheet.cell(row=y,column=6).value = 'Вес'
  sheet.cell(row=y,column=7).value = 'Вес общий'
  for i in stat:
    y+=1
    sheet.cell(row=y,column=1).value = i.detail
    sheet.cell(row=y,column=2).value = i.point.assembly.assembly
    sheet.cell(row=y,column=3).value = i.point.name
    sheet.cell(row=y,column=4).value = i.point.draw
    sheet.cell(row=y,column=5).value = int(i.count)
    sheet.cell(row=y,column=6).value = i.point.assembly.weight
    sheet.cell(row=y,column=7).value = i.point.assembly.weight * i.count
    sheet.cell(row=y,column=6).number_format = BUILTIN_FORMATS[1]
    sheet.cell(row=y,column=7).number_format = BUILTIN_FORMATS[1]
  book.save('stat '+str(faza)+' '+case+'.xlsx')

  # stat = Point.select(Point,Drawing,fn.SUM(Drawing.weight).alias('weight'),fn.COUNT(Point.assembly).alias('count'),(fn.SUM(Drawing.weight) * fn.COUNT(Point.assembly)).alias('weight_')).join(Drawing).join(Order).where(Order.cas == case).group_by(Point.assembly).order_by(Drawing.assembly)
  # y = 1
  # sheet.cell(row=y,column=1).value = 'Марка'
  # sheet.cell(row=y,column=2).value = 'Количество'
  # sheet.cell(row=y,column=3).value = 'Вес'
  # sheet.cell(row=y,column=4).value = 'Вес общий'
  # for i in stat:
  #   y+=1
  #   sheet.cell(row=y,column=1).value = i.assembly.assembly
  #   sheet.cell(row=y,column=2).value = int(i.count)
  #   sheet.cell(row=y,column=3).value = i.assembly.weight
  #   sheet.cell(row=y,column=4).value = i.assembly.weight * i.count
  # book.save('statall '+str(faza)+' '+case+'.xlsx')


def Cuting(faza,case):
  book = Workbook()
  sheet = book.active
  cut = PointPart.select(Part.number,Part.size,Point.draw,fn.SUM(Part.count).alias('count')).join(Point).join(Drawing).join(Order).join_from(PointPart,Part).where(Order.cas == case,Point.faza == faza,Part.profile == 'Лист').group_by(PointPart.part).order_by(Part.number,Point.draw)
  y = 1
  sheet.cell(row=y,column=1).value = 'Деталь'
  sheet.cell(row=y,column=2).value = 'Чертеж'
  sheet.cell(row=y,column=3).value = 'Толщина'
  sheet.cell(row=y,column=4).value = 'Количество'
  for i in cut:
    y+=1
    sheet.cell(row=y,column=1).value = i.part.number
    sheet.cell(row=y,column=2).value = i.point.draw
    sheet.cell(row=y,column=3).value = i.part.size
    sheet.cell(row=y,column=4).value = i.count
  book.save('cut '+str(faza)+' '+case+'.xlsx')

def User_worker():
  wb = load_workbook('VVL.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('users')
  # d = []
  post = []
  for y in range(3,39):
    b = (sheet.cell(row=y,column=1).value).replace('.',' ').strip().split(' ')
    if len(b) == 2:
      b.append(' ')
    user = User.create(surname=b[0],name=b[1],patronymic=b[2])
    for i in range(3,13):
      if sheet.cell(row=y,column=i).value == 1:
        post.append([user,sheet.cell(row=2,column=i).value,sheet.cell(row=1,column=i).value])
  Worker.insert_many(post, fields=[Worker.user,Worker.oper,Worker.oper_rus]).execute()