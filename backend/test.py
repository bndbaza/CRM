from models import *
from datetime import datetime, timedelta
import os
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.styles import PatternFill as ft

from peewee import fn, Case
from package_list import PackList
from tekla import Size, Test

def AAA():
  import csv

  tekla = {}
  sql = {}
  with open('2325.16.xls','r', encoding='windows-1251',newline='') as file:
    reader = csv.reader(file, delimiter='\t')
    for row in reader:
      if row[3].replace(' ','') == '1' and row[0].replace(' ','') == 'DRAWING' and row[1].replace(' ','').find('(?)') == -1:
        tekla[row[1].replace(' ','')] = {
          'area':float(row[4].replace(' ','')),
          'count':int(row[5].replace(' ','')),
          'part':{},
          'point':[],
          'weld':[],
          'bolt':[],
          'nut':[],
          'washer':[]
        }
      if row[0].replace(' ','') == 'ASSEMBLY' and row[1].replace(' ','') in tekla.keys():
        try:
          if row[7].replace(' ','') == '':
            draw = None
          else:
            draw = int(row[7].replace(' ',''))
        except:
          draw = None
        tekla[row[1].replace(' ','')]['point'].append({
          'point_x':row[3].replace(' ','').replace('<','').replace('>','').split('/')[0].split('-')[0],
          'point_y':row[3].replace(' ','').replace('<','').replace('>','').split('/')[-1].split('-')[0],
          'point_z':float(row[2].replace(' ','')),
          'name':row[5].strip(),
          'draw':draw
        })
      if row[0].replace(' ','') == 'PART' and row[1].replace(' ','') in tekla.keys():
        profile = Size(row[4].replace(' ',''))
        tekla[row[1].replace(' ','')]['part'][int(row[2].replace(' ',''))] = {
          'count':int(row[3].replace(' ','')),
          'profile':profile[0],
          'size':profile[1],
          'length':float(row[5].replace(' ','')),
          'weight':float(row[6].replace(' ','')),
          'mark':row[8].replace(' ',''),
          'manipulation':row[9].replace(' ',''),
          'work':profile[2],
          'width':profile[3],
          'perimeter':int(Test(row[11].replace(' ',''))),
          'depth':float(Test(row[12].replace(' ',''),profile[1])),
          'hole':[],
        }
      if row[0].replace(' ','') == 'WEIGHT' and row[1].replace(' ','') in tekla.keys():
        if row[3].replace(' ','') != '':
          more = 0
          more = float(row[3].replace(' ','').split(';')[1])
        else:
          more = 0
        tekla[row[1].replace(' ','')]['weight'] = {
          'weight':float(row[2].replace(' ','')),
          'more':more,
        }
      if row[0].replace(' ','') == 'WELD' and row[1].replace(' ','') in tekla.keys():
        tekla[row[1].replace(' ','')]['weld'].append({
          'cathet':int(float(row[2].replace(' ',''))),
          'length':float(row[3].replace(' ','')) / tekla[row[1].replace(' ','')]['count'],
          'count':int(row[4].replace(' ','')) / tekla[row[1].replace(' ','')]['count']
        })
      if row[0].replace(' ','') == 'BOLT' and row[1].replace(' ','') in tekla.keys():
        tekla[row[1].replace(' ','')]['bolt'].append({
          'profile':row[2].replace(' ',''),
          'gost':row[3].replace(' ',''),
          'count':int(row[4].replace(' ','')),
          'weight':float(row[5].replace(' ',''))
        })
      if row[0].replace(' ','') == 'NUT' and row[1].replace(' ','') in tekla.keys():
        tekla[row[1].replace(' ','')]['nut'].append({
          'profile':row[2].replace(' ',''),
          'gost':row[3].replace(' ',''),
          'count':int(row[4].replace(' ','')),
          'weight':float(row[5].replace(' ',''))
        })
      if row[0].replace(' ','') == 'WASHER' and row[1].replace(' ','') in tekla.keys():
        tekla[row[1].replace(' ','')]['washer'].append({
          'profile':row[2].replace(' ',''),
          'gost':row[3].replace(' ',''),
          'count':int(row[4].replace(' ','')),
          'weight':float(row[5].replace(' ',''))
        })
      # if row[0].replace(' ','') == 'HOLE' and row[1].replace(' ','') in tekla.keys():
      #   if tekla[row[1].replace(' ','')]['part'][int(row[2].replace(' ',''))]['profile'] == 'Лист' and int(tekla[row[1].replace(' ','')]['part'][int(row[2].replace(' ',''))]['size']) < 15:
      #         pass
      #   else:
      #     tekla[row[1].replace(' ','')]['part'][int(row[2].replace(' ',''))]['hole'].append({
      #       'diameter':int(row[3].replace(' ','')),
      #       'count':(int(row[4].replace(' ','')))/tekla[row[1].replace(' ','')]['part'][int(row[2].replace(' ',''))]['count'],
      #       'depth':int(row[5].replace(' ','')) / 2
      #     })
      # if row[0].replace(' ','') == 'CHAMFER' and row[1].replace(' ','') in tekla.keys():
      #   tekla[row[1].replace(' ','')]['part'][int(row[2].replace(' ',''))]['chamfer'] = {
      #     'length':float(row[3].replace(' ',''))
      #   }


  for row in Drawing.select().where(Drawing.cas == 15):
    sql[row.assembly] = {
      'area':float(row.area),
      'count':row.count,
      'part':{},
      'point':[],
      'weld':[],
      'bolt':[],
      'nut':[],
      'washer':[],
      'weight':{'weight':float(row.weight),'more':float(row.more)}
    }
    for point in row.points:
      sql[row.assembly]['point'].append({
        'point_x':point.point_x,
        'point_y':point.point_y,
        'point_z':float(point.point_z),
        'name':point.name,
        'draw':point.draw
      })
    for weld in row.welds:
      sql[row.assembly]['weld'].append({
        'cathet':weld.cathet,
        'length':int(weld.length),
        'count':weld.count
      })
    for bolt in row.bolts:
      sql[row.assembly]['bolt'].append({
        'profile':bolt.profile,
        'gost':bolt.gost,
        'count':int(bolt.count),
        'weight':float(bolt.weight)
      })
    for nut in row.nuts:
      sql[row.assembly]['nut'].append({
        'profile':nut.profile,
        'gost':nut.gost,
        'count':int(nut.count),
        'weight':float(nut.weight)
      })
    for washer in row.washers:
      sql[row.assembly]['washer'].append({
        'profile':washer.profile,
        'gost':washer.gost,
        'count':int(washer.count),
        'weight':float(washer.weight)
      })
    for part in row.parts:
      if part.number in sql[row.assembly]['part'].keys():
        sql[row.assembly]['part'][str(part.number)+' '+str(datetime.now())] = {
          'count':int(part.count),
          'profile':part.profile,
          'size':part.size,
          'length':float(part.length),
          'weight':float(part.weight),
          'mark':part.mark,
          'manipulation':part.manipulation,
          'work':part.work,
          'width':part.width,
          'perimeter':int(part.perimeter),
          'depth':float(part.depth),
          'hole':[],
        }
      else:
        sql[row.assembly]['part'][part.number] = {
          'count':int(part.count),
          'profile':part.profile,
          'size':part.size,
          'length':float(part.length),
          'weight':float(part.weight),
          'mark':part.mark,
          'manipulation':part.manipulation,
          'work':part.work,
          'width':part.width,
          'perimeter':int(part.perimeter),
          'depth':float(part.depth),
          'hole':[],
        }
      # for hole in part.holes:
      #   sql[row.assembly]['part'][part.number]['hole'].append({
      #     'diameter':int(hole.diameter),
      #     'count':hole.count,
      #     'depth':hole.depth
      #   })
      # for chamfer in part.chamfers:
      #   sql[row.assembly]['part'][part.number]['chamfer'] = {
      #     'length':float(chamfer.length)
      #   }
  

  delta = [
    ('count','Количество'),
    ('profile','Профиль'),
    ('size','Размер'),
    ('length','Длина'),
    ('weight','Вес'),
    ('mark','Марка'),
    ('manipulation','Операции'),
    ('work','В работе'),
    ('width','Ширина'),
    ('perimeter','Периметер'),
    ('depth','Глубина'),
  ]
  corr = {'delete':[],'update':[],'error':[],'create':[]}
  for s in sql.keys():
    try:
      tekla[s]['part']
    except:
      corr['error'].append(s)
      break

    if sql[s]['part'] != tekla[s]['part']:
      aa = sql[s]['part'].copy()
      bb = tekla[s]['part'].copy()
      for a in sql[s]['part']:
        try:
          if aa[a] == bb[a]:
            del aa[a]
            del bb[a]
        except:
          pass
      for a in aa:
        if bb.get(a) == None:
          corr['delete'].append((s,a))
          # pr = Part.select().join(Drawing).where(Drawing.assembly == s,Part.number == a).first()
          # Hole.delete().where(Hole.part == pr).execute()
          # Chamfer.delete().where(Chamfer.part == pr).execute()
          # PointPart.delete().where(PointPart.part == pr).execute()
          # TaskPart.delete().where(TaskPart.part == pr).execute()
          # Part.delete().where(Part.id == pr).execute()
        else:
          corr['update'].append((s,a,aa.get(a),bb.get(a)))
          # pr = Part.select().join(Drawing).where(Drawing.assembly == s,Part.number == a).first()
          # pr.count = bb[a]['count']
          # pr.profile = bb[a]['profile']
          # pr.size = bb[a]['size']
          # pr.length = bb[a]['length']
          # pr.weight = bb[a]['weight']
          # pr.mark = bb[a]['mark']
          # pr.manipulation = bb[a]['manipulation']
          # pr.work = bb[a]['work']
          # pr.width = bb[a]['width']
          # pr.perimeter = bb[a]['perimeter']
          # pr.depth = bb[a]['depth']
          # pr.save()
          del bb[a]
      if bb != {}:
        corr['create'].append((s,bb))
  
  for i in corr['delete']:
    print(f'Удалена деталь № {i[1]} марки {i[0]}')
  for i in corr['update']:
    for d in delta:
      if i[2][d[0]] != i[3][d[0]]:
        print(f'Марка {i[0]} деталь {i[1]} {d[1]} В базе: {i[2][d[0]]}, В файле: {i[3][d[0]]}')
  for i in corr['create']:
    for y in i[1]:
      print(f'Добавлена новая деталь № {y} к марке {i[0]}')
      
  dates = datetime.today()
  create = []
  for t in tekla:
    # if tekla[t]['weld'] != sql[t]['weld']:
    delete = Drawing.get(Drawing.assembly == t)
      # Weld.delete().where(Weld.assembly == delete.id).execute()
    for y in tekla[t]['weld']:
      create.append((delete,y['cathet'],y['length'],y['count'],dates))
  with connection.atomic():
    Weld.insert_many(create, fields=[Weld.assembly,Weld.cathet,Weld.length,Weld.count,Weld.create_date]).execute()
  


def BBB():
  PrintBirk.delete().execute()
  wb = load_workbook('paint.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Лист1')
  post = []
  for y in range(1,302):
    post.append(sheet.cell(row=y,column=1).value)

  detail = PointPart.select(PointPart,fn.MAX(Part.length).alias('length')).join(Part).join_from(PointPart,Point).where(PointPart.detail.in_(post)).group_by(PointPart.detail)
  pr = []
  for d in detail:
    pr.append((d.detail,d.length,d.part.assembly.weight,d.part.assembly.cas.cas,d.point.faza,d.part.assembly.assembly,'Run '+str(d.detail)+' weld '+d.part.assembly.cas.cas,d.point.draw,2))
  PrintBirk.insert_many(pr, fields=[PrintBirk.detail,PrintBirk.length,PrintBirk.weight,PrintBirk.case,PrintBirk.faza,PrintBirk.mark,PrintBirk.qr,PrintBirk.draw,PrintBirk.count]).execute()

  z = PrintBirk.select()

  book = Workbook()
  sheet = book.active
  y = 1
  sheet.cell(row=y,column=1).value = 'наряд'
  sheet.cell(row=y,column=2).value = 'длина'
  sheet.cell(row=y,column=3).value = 'вес'
  sheet.cell(row=y,column=4).value = 'заказ'
  sheet.cell(row=y,column=5).value = 'фаза'
  sheet.cell(row=y,column=6).value = 'марка'
  sheet.cell(row=y,column=7).value = 'qr'
  sheet.cell(row=y,column=8).value = 'чертёж'
  sheet.cell(row=y,column=9).value = 'Количество'
  for i in z:
    y += 1
    sheet.cell(row=y,column=1).value = i.detail
    sheet.cell(row=y,column=2).value = i.length
    sheet.cell(row=y,column=3).value = float(i.weight)
    sheet.cell(row=y,column=4).value = i.case
    sheet.cell(row=y,column=5).value = i.faza
    sheet.cell(row=y,column=6).value = i.mark
    sheet.cell(row=y,column=7).value = i.qr
    sheet.cell(row=y,column=8).value = i.draw
    sheet.cell(row=y,column=9).value = i.count
  book.save('Пример.xlsx')





def DDD():
  # d = [22,23,24,25,26]
  # faza = Faza.select().where(Faza.faza.in_(d))
  # for i in faza:
  #   pp = PointPart.select(PointPart,fn.SUM(Drawing.weight).alias('weight'),fn.SUM(PointPart.weld).alias('sum_weld')).join(Point).join(Drawing).group_by(PointPart.detail).having(PointPart.detail == i.detail).first()
  #   if pp.sum_weld == 0:
  #     x = 1
  #     print(x)
  #   else:
  #     x = pp.sum_weld
  #   i.weight = pp.weight / x
  # Faza.bulk_update(faza,fields=[Faza.weight])

  faza = Faza.select().where(Faza.assembly != 0,Faza.set != 3)
  for i in faza:
    print(i.detail)
  #   i.set = 3
  # with connection.atomic():
  #   Faza.bulk_update(faza,fields=[Faza.set])
  print('-----')

  faza = Faza.select().where(Faza.paint != 0,Faza.weld != 3)
  for i in faza:
    print(i.detail)
  #   i.paint = 0
  # with connection.atomic():
  #   Faza.bulk_update(faza,fields=[Faza.paint])
  



def GGG():
  # z = Faza.select()
  # for i in z:
  #   i.kmd = 3
  #   i.preparation = 1
  #   i.in_work = 1
  # with connection.atomic():
  #   Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work])
  
  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'paint',Detail.end != None).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 3
      i.assembly = 3
      i.weld = 3
      i.paint = 3
      i.shipment = 1
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint,Faza.shipment])
    print(1)
  except:
    print('error')
    print(1)

  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'paint',Detail.end == None,Detail.start != None).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 3
      i.assembly = 3
      i.weld = 3
      i.paint = 2
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint])
    print(2)
  except:
    print('error')
    print(2)

  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'paint',Detail.start == None,Detail.to_work == 1).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 3
      i.assembly = 3
      i.weld = 3
      i.paint = 1
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint])
    print(3)
  except:
    print('error')
    print(3)

  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'weld',Detail.end != None).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 3
      i.assembly = 3
      i.weld = 3
      # i.paint = 3
      # i.shipment = 1
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint,Faza.shipment])
    print(4)
  except:
    print('error')
    print(4)

  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'weld',Detail.end == None,Detail.start != None).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 3
      i.assembly = 3
      i.weld = 2
      # i.paint = 2
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint])
    print(5)
  except:
    print('error')
    print(5)

  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'weld',Detail.start == None,Detail.to_work == 1).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 3
      i.assembly = 3
      i.weld = 1
      # i.paint = 1
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint])
    print(6)
  except:
    print('error')
    print(6)

  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'assembly',Detail.end != None).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 3
      i.assembly = 3
      # i.weld = 3
      # i.paint = 3
      # i.shipment = 1
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint,Faza.shipment])
    print(7)
  except:
    print('error')
    print(7)

  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'assembly',Detail.end == None,Detail.start != None).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 3
      i.assembly = 2
      # i.weld = 2
      # i.paint = 2
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint])
    print(8)
  except:
    print('error')
    print(8)

  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'assembly',Detail.start == None,Detail.to_work == 1).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 3
      i.assembly = 1
      # i.weld = 1
      # i.paint = 1
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint])
    print(9)
  except:
    print('error')
    print(9)

  
  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'set',Detail.end != None).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 3
      # i.assembly = 3
      # i.weld = 3
      # i.paint = 3
      # i.shipment = 1
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint,Faza.shipment])
    print(10)
  except:
    print('error')
    print(10)

  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'set',Detail.end == None,Detail.start != None).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 2
      # i.assembly = 2
      # i.weld = 2
      # i.paint = 2
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint])
    print(11)
  except:
    print('error')
    print(11)

  try:
    d = Detail.select(Detail.detail).where(Detail.oper == 'set',Detail.start == None,Detail.to_work == 1).tuples()
    z = Faza.select().where(Faza.detail.in_(d))
    for i in z:
      i.kmd = 3
      i.preparation = 3
      i.in_work = 3
      i.set = 1
      # i.assembly = 1
      # i.weld = 1
      # i.paint = 1
    with connection.atomic():
      Faza.bulk_update(z,fields=[Faza.kmd,Faza.preparation,Faza.in_work,Faza.set,Faza.assembly,Faza.weld,Faza.paint])
    print(12)
  except:
    print('error')
    print(12)

  # for i in range(1,22):
  #   taskpart = TaskPart.select(TaskPart.part).join(Task).where(Task.faza == i,Task.start != None).tuples()
  #   pointpart = PointPart.select(PointPart.detail).join(Point).where(Point.faza == i,PointPart.part.in_(taskpart)).group_by(PointPart.detail).tuples()
  #   faza = Faza.select().where(Faza.preparation != 3,Faza.detail.in_(pointpart))
  #   if len(faza) != 0:
  #     for y in faza:
  #       y.preparation = 2
  #     with connection.atomic():
  #       Faza.bulk_update(faza,fields=[Faza.preparation])

  # task = Task.select(Task.faza).where(Task.start != None).group_by(Task.faza).tuples()
  # z = Faza.select().where(Faza.in_work == 1,Faza.faza.in_(task))
  # for i in z:
  #   i.in_work = 3
  # with connection.atomic():
  #   Faza.bulk_update(z,fields=[Faza.in_work])



def CCC():
  details = Detail.select().where(Detail.worker_1 != None)
  data = []
  for detail in details:
    if detail.worker_2 != None:
      data.append((detail.id,detail.worker_1,detail.faza.weight / 2,detail.norm / 2))
      data.append((detail.id,detail.worker_2,detail.faza.weight / 2,detail.norm / 2))
    else:
      data.append((detail.id,detail.worker_1,detail.faza.weight,detail.norm))
  with connection.atomic():
    DetailUser.insert_many(data,fields=[DetailUser.detail,DetailUser.worker,DetailUser.weight,DetailUser.norm]).execute()

def XXX():
  data = Detail.select(Detail.detail).where(Detail.oper == 'assembly',Detail.norm == 0).tuples()
  d = {}
  for i in PointPart.select(PointPart,fn.SUM(Part.count).alias('count')).join(Part).where(PointPart.detail.in_(data)).group_by(PointPart.detail):

    norm_assembly = AssemblyNorm.select().where(AssemblyNorm.name == i.point.name,
                                                i.point.assembly.weight >= AssemblyNorm.mass_of,
                                                i.point.assembly.weight < AssemblyNorm.mass_to,
                                                i.count >= AssemblyNorm.count_of,
                                                i.count < AssemblyNorm.count_to).first()
    # print(i.detail,norm_assembly.norm  * (i.point.assembly.weight / 1000))
    d[i.detail] = norm_assembly.norm  * (i.point.assembly.weight / 1000)
  data2 = Detail.select().where(Detail.oper == 'assembly',Detail.norm == 0)
  for y in data2:
    y.norm = d[y.detail]
  with connection.atomic():
    Detail.bulk_update(data2,fields=[Detail.norm])


def HHH():

  # drawing = Drawing.select()
  # for i in drawing:
  #   if i.count != len(i.points):
  #     print(i.assembly,i.count,len(i.points))
  # point = Point.select().join(Drawing).where(Drawing.assembly == 'МШ-1').group_by(Point.faza)
  # for i in point:
  #   print(i.faza)

  # pp = PointPart.select(PointPart.part).join(Point).join(Drawing).where(Drawing.assembly == 'МК-1',Point.faza == 19).tuples()

  # pp = Point.select().join(Drawing).where(Drawing.assembly == 'МК-1',Point.faza == 19).first()





  dr = Drawing.get(Drawing.assembly == 'МБ2-144')
  det = 952
  f = 6
  tp = TaskPart.select(TaskPart.id).join(Task).join_from(TaskPart,Part).where(Task.faza == f,Part.assembly == dr).tuples()
  TaskPart.delete().where(TaskPart.id.in_(tp)).execute()


  # tp = TaskPart.select().join(Task).join_from(TaskPart,Part).where(Task.faza == f,Part.assembly == dr)
  # for t in tp:
  #   t.count = t.count / 2
  #   t.save()
  # ppp = PointPart.select().where(PointPart.detail == det).first()

  d = Detail.select(Detail.id).where(Detail.detail == det).tuples()
  DetailUser.delete().where(DetailUser.detail.in_(d)).execute()
  Detail.delete().where(Detail.detail == det).execute()
  Faza.delete().where(Faza.detail == det).execute()
  PointPart.delete().where(PointPart.detail == det).execute()
  Point.delete().where(Point.assembly == dr,Point.faza == f).execute()


  # pp = Point.select().where(Point.assembly == dr)
  # for p in pp:
  #   print(p,p.draw,p.faza,p.point_z,p.pointparts[0].detail)

  # tp = TaskPart.select().join(Task).join_from(TaskPart,Part).where(Task.faza == f,Part.assembly == dr)
  # for t in tp:
  #   print(t,t.count)




  # print(len(point),point[0].assembly.count)

  # drawing = Drawing.select()
  # for i in drawing:
  #   if i.count != len(i.points):
  #     print(i.assembly,i.count,len(i.points))
  # for i in point:
  #   print(i.assembly.assembly,i.faza,i.point_x,i.point_y,i.point_z,i.draw)

  # for i in Detail.select().where(Detail.oper == 'weld',Detail.worker_2 != None):
  #   print(i.detail,i.end)

  # faza = Faza.select().where(Faza.paint == 3,Faza.shipment == 0)
  # print(len(faza))
  # for i in faza:
  #   i.shipment = 1
  # with connection.atomic():
  #   Faza.bulk_update(faza,fields=[Faza.shipment])

  # start = datetime(2022,4,1)
  # end = datetime(2022,5,1)
  # d = DetailUser.select(Detail.detail).join(Detail).where(DetailUser.worker == 2,Detail.end >= start,Detail.end < end).tuples()
  # pp = PointPart.select(PointPart.point).where(PointPart.detail.in_(d)).group_by(PointPart.point).tuples()
  # weld = Weld.select(Weld,fn.SUM(Weld.length)).join(Drawing).join(Point).where(Point.id.in_(pp)).group_by(Weld.cathet)
  # for i in weld:
    # cathet = WeldNorm.get(WeldNorm.cathet == i.cathet)
    # x = i.length / 1000 * cathet.norm
    # print(i.assembly.assembly,i.cathet,i.length)
    # inde/x += x
  # print(index / 60)

  # pp = PointPart.select().join(Point).join(Drawing).join(Weld).where(PointPart.detail.in_(d))#.group_by(Point.id)
  # for i in pp:
  #   print(i.point.assembly)

  # for i in Weld.select(Weld,fn.SUM(Weld.length).alias('len')).where(Weld.assembly.in_(pp)).group_by(Weld.cathet):
  #   print(i.cathet,i.len)





  # detail = Detail.select().where(Detail.oper == 'paint',Detail.start != None,Detail.end == None)
  # for i in detail:
  #   print(i.detail)
  #   i.end = i.start
  # with connection.atomic():
  #   Detail.bulk_update(detail,fields=[Detail.end])

  # faza = Faza.select().where(Faza.paint == 2)
  # for y in faza:
  #   print(y.detail)
  #   y.paint = 3
  # with connection.atomic():
  #   Faza.bulk_update(faza,fields=[Faza.paint])
  
  # for i in range(1,22):
  #   # PointPart.select().join(Point).where()
  #   # detail = Detail.select().where(Detail.start == None)
  #   faza = Faza.select().where(Faza.preparation == 0,Faza.faza == i,Faza.case == 15)
  #   print(i,len(faza))


  # faza = Faza.select(Faza.detail).where(Faza.preparation == 2).tuples()
  # for i in faza:
  #   detail = Detail.get(Detail.detail == i,Detail.oper == 'set')
  #   print(i[0],detail.start,detail.end,detail.to_work)


  # for index in range(1,22):
  #   y = PointPart.select(PointPart.detail).join(Point).join(Drawing).where(Point.faza == index,Drawing.cas == 15).group_by(PointPart.detail).tuples()
  #   z = Detail.select().where(Detail.detail.in_(y))
  #   for i in z:
  #     i.faza = index
  #     i.case = 15
  #   with connection.atomic():
  #     Detail.bulk_update(z,fields=[Detail.faza,Detail.case])
  # y = Part.select().where(Part.assembly == 1407)
  # z = Drawing.select().where(Drawing.id == 1407).first()
  # print(z.create_date)
  # for i in y:
    # i.create_date = z.create_date
    # i.save()
    # print(i.create_date)

  # point = Point.select(Point.id).join(Drawing).where(Drawing.assembly == 'МБ2-233')
  # for i in point:
  #   print(i)
  # part = Part.select().join(Drawing).where(Drawing.assembly == 'МБ2-233')
  # for i in part:
  #   cgm = 1
  #   saw = 0
  #   if i.profile != 'Лист':
  #     cgm = 0
  #     saw = 1
  #   PointPart.create(point = 4867,part = i,detail = 1361,cgm = cgm,saw = saw)


  # r = []
  # for i in PointPart.select(PointPart,fn.SUM(Drawing.weight).alias('weight'),fn.SUM(PointPart.weld).alias('sum_weld')).join(Point).join(Drawing).group_by(PointPart.detail):
  #   if i.sum_weld == 0:
  #     x = 1
  #   else:
  #     x = i.sum_weld
  #   r.append((i.detail,i.point.faza,i.point.assembly.cas,i.weight / x))
  # with connection.atomic():
  #   Faza.insert_many(r, fields=[Faza.detail,Faza.faza,Faza.case,Faza.weight]).execute()



  # z = Detail.select().where(Detail.faza == None)
  # for i in z:
  #   y = Faza.select().where(Faza.detail == i.detail).first()
  #   i.faza = y
  # with connection.atomic():
  #   Detail.bulk_update(z,fields=[Detail.faza])


  # for report in Detail.select(
  #   Detail,
  #   fn.SUM(Case(None,[(Detail.oper == 'paint',Faza.weight)],0)).alias('weight_create'),
  #   fn.SUM(Case(None,[(Detail.oper == 'paint',Faza.weight)],0)).alias('weight_in_work'),
  #   fn.SUM(Case(None,[((Detail.oper == 'paint') & (Detail.to_work == 1),Faza.weight)],0)).alias('weight_in_preparation'),
  #   fn.SUM(Case(None,[((Detail.oper == 'paint') & (Detail.to_work == 1),Faza.weight)],0)).alias('weight_set')
  # ).join(Faza).group_by(Faza.faza):
  #   print(
  #     report.faza.faza,
  #     report.weight_create / 1000,
  #     report.weight_in_work / 1000,
  #     report.weight_in_preparation / 1000,
  #   )

  # z = Faza.select().where(Faza.preparation == 0)
  # for i in z:
  #   i.preparation = 1
  #   i.in_work = 1
  # with connection.atomic():
  #   Faza.bulk_update(z,fields=[Faza.preparation,Faza.in_work])




  # task = Task.select(Task.task).where(Task.start != None).tuples()
  # taskfaza = Task.select(Task.task).where(Task.start != None).tuples()
  # taskpart = TaskPart.select().where(TaskPart.task.in_(task))
  # pointpart = PointPart.select()
  # z = Faza.select().where(Faza.in_work == 1,Faza.faza.in_(task))
  # for i in z:
  #   i.in_work = 3
  # with connection.atomic():
  #   Faza.bulk_update(z,fields=[Faza.in_work])

  # d = (19,20,21)
  # pp = PointPart.select(PointPart.detail).join(Point).where(Point.faza.in_(d)).group_by(PointPart.detail).tuples()
  # Detail.delete().where(Detail.detail.in_(pp)).execute()

  # task = Task.select(Task.id).where(Task.faza.in_(d)).tuples()
  # TaskPart.delete().where(TaskPart.task.in_(task)).execute()
  # Task.delete().where(Task.faza.in_(d)).execute()

  # dpp = PointPart.select(PointPart.id).join(Point).where(Point.faza.in_(d)).tuples()
  # PointPart.delete().where(PointPart.id.in_(dpp)).execute()

def YYY():
  drawing = Drawing.get(Drawing.assembly == 'МБ1-4')
  part = Part.select().where(Part.assembly == drawing)
  partid = Part.select(Part.id).where(Part.assembly == drawing).tuples()

  Chamfer.delete().where(Chamfer.part.in_(part)).execute()
  Hole.delete().where(Hole.part.in_(part)).execute()
  Part.delete().where(Part.id.in_(partid)).execute()

  Point.delete().where(Point.assembly == drawing).execute()
  Weld.delete().where(Weld.assembly == drawing).execute()
  Bolt.delete().where(Bolt.assembly == drawing).execute()
  Nut.delete().where(Nut.assembly == drawing).execute()
  Washer.delete().where(Washer.assembly == drawing).execute()
  Drawing.delete().where(Drawing.id == drawing.id).execute()



def ZZZ():
  # pp = PointPart.select().join(Point).join(Drawing).where(Drawing.cas == 20).group_by(PointPart.detail)
  # for i in pp:
  #   print(i.detail,i.point.assembly.assembly)

  # PointPart.update({PointPart.detail:2341}).where(PointPart.detail == 2359).execute()
  # PointPart.update({PointPart.detail:2342}).where(PointPart.detail == 2360).execute()
  # PointPart.update({PointPart.detail:2343}).where(PointPart.detail == 2361).execute()
  # PointPart.update({PointPart.detail:2344}).where(PointPart.detail == 2362).execute()
  # PointPart.update({PointPart.detail:2345}).where(PointPart.detail == 2363).execute()
  # PointPart.update({PointPart.detail:2346}).where(PointPart.detail == 2364).execute()
  # PointPart.update({PointPart.detail:2347}).where(PointPart.detail == 2365).execute()
  # PointPart.update({PointPart.detail:2348}).where(PointPart.detail == 2366).execute()
  # PointPart.update({PointPart.detail:2349}).where(PointPart.detail == 2367).execute()

  PointPart.update({PointPart.detail:2350}).where(PointPart.detail == 2359).execute()
  PointPart.update({PointPart.detail:2351}).where(PointPart.detail == 2360).execute()
  PointPart.update({PointPart.detail:2352}).where(PointPart.detail == 2361).execute()


def UUU():
  book = Workbook()
  sheet_n = book.active
  index = 1
  wb = load_workbook('mark2.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Лист1')
  post = []
  for y in range(1,302):
    #post.append(sheet.cell(row=y,column=1).value)
    sheet_n.cell(row=index,column=1).value = sheet.cell(row=y,column=2).value
    sheet_n.cell(row=index,column=2).value = sheet.cell(row=y,column=3).value
    pp = PointPart.select(PointPart.detail).join(Point).join(Drawing).join(Order).where(Drawing.assembly == sheet.cell(row=y,column=2).value,Order.cas == '2325').group_by(PointPart.detail).tuples()
    faza = Faza.select().where(Faza.detail.in_(pp))
    result = ''
    status = ''
    for stage in faza:
      if stage.set == 0:
        status = 'Заготовка'
      elif stage.set == 1:
        status = 'Готов к комплектации'
      elif stage.set == 2:
        status = 'В комплектации'
      elif stage.assembly == 1:
        status = 'Готов к сборке'
      elif stage.assembly == 2:
        status = 'В сборке'
      elif stage.weld == 1:
        status = 'Готов к сварке'
      elif stage.weld == 2:
        status = 'В сварке'
      elif stage.paint == 1:
        status = 'Готов к покраске'
      elif stage.paint == 2:
        status = 'В покраске'
      elif stage.packed == 1:
        status = 'Готов к упаковке'
      elif stage.packed == 3:
        status = 'Упаковано'
      else:
        status = '?'
      sheet_n.cell(row=index,column=3).value = stage.detail
      sheet_n.cell(row=index,column=4).value = status
      index += 1
  book.save(f'222.xlsx')

def USS():
  # pp = PointPart.select().join(Point).join(Drawing).where(Drawing.assembly == 'МСг-10')
  # for i in pp:
  #   print(i.detail)
  # p = Point.select().join(Drawing).where(Drawing.assembly.in_(['МСг-10','МСг-11','МСг-12','МСг-13']))
  # for i in p:
  #   i.name = 'Связь горизонтальная'
  #   i.save()
  #   print(i.faza,i.name)

  # part = Part.select(fn.SUM(Part.weight * Part.count)).join(Drawing).where(Drawing.cas == 20).group_by(Drawing.cas).scalar()
  # part = Part.select(fn.SUM(Part.weight)).join(Drawing).where(Drawing.cas == 21).group_by(Drawing.cas).scalar()
  # part = Part.select().join(Drawing).where(Drawing.cas == 21)
  # print(part)
  # for i in part:
  #   print(i,i.sn,i.count,i.weight)

  # faza = Faza.select(fn.COUNT(Faza.paint)).where(Faza.paint == 3,Faza.packed == 0).group_by(Faza.paint).scalar()
  # detail = Detail.select().join(Faza).where(Detail.oper == 'paint',Detail.to_work == 1,Faza.paint == 0)
  # for d in detail:
  #   # d.to_work = 0
  #   # d.save()
  #   print(d.detail)
  # pp = PointPart.update({PointPart.turning: 1}).where(PointPart.part == 8441).execute()
  # pp = PointPart.select().where(PointPart.part == 8441)
  # for i in pp:
  #   print(i)

  # part = Part.select().join(Drawing).where(Drawing.cas == 20)
  # print(part)
  # book = Workbook()
  # sheet = book.active
  # y = 1
  # for i in part:
  #   print(i.sn,i.number,i.weight)

  #   sheet.cell(row=y,column=1).value = i.sn
  #   sheet.cell(row=y,column=2).value = i.number
  #   sheet.cell(row=y,column=3).value = i.weight
  #   y+=1


  # book.save(f'999.xlsx')



  # wb = load_workbook('999.xlsx',data_only=True)
  # sheet = wb.get_sheet_by_name('Sheet')
  # for y in range(1,82):
  #   part = Part.select().join(Drawing).where(
  #     Drawing.cas == 20,
  #     Part.number == sheet.cell(row=y,column=2).value,
  #     Part.sn ==sheet.cell(row=y,column=1).value
  #   ).first()
  #   part.weight = sheet.cell(row=y,column=3).value
  #   part.save()

  # point = Drawing.select().where(Drawing.cas == 20,Drawing.assembly == 'ёмкость').first()
  # print(point,point.weight)

  # parts = Part.select(Part,fn.SUM(Part.weight * Part.count).alias('aaa')).join(Drawing).where(Drawing.cas == 20).group_by(Part.assembly)
  # for part in parts:
  #   faza = Faza.get()
  #   dr = part.assembly
  #   dr.weight = part.aaa
  #   dr.save()
  #   print(part.aaa,part.assembly)

  # pp = PointPart.select().join(Part).join(Drawing).where(Drawing.cas == 20).group_by(PointPart.detail)
  # for p in pp:
  #   faza = Faza.get(Faza.detail == p.detail)
  #   print(faza.weight,p.part.assembly.weight)
  #   faza.weight = p.part.assembly.weight
  #   faza.save()

  # point = Point.select(Point,fn.SUM(Drawing.weight).alias('we')).join(Drawing).group_by(Point.name)
  # for p in point:
  #   print(p.name,p.we)

  # for i in pp:
  #   print(len(pp))


  cor = 2678

  index = PointPart.select(fn.MAX(PointPart.detail)).scalar() +1


  for i in range(64):
    pp = PointPart.select().where(PointPart.detail == cor)
    if len(pp) == 1:
      break
    pp = pp.first()
    pp.detail = index
    pp.save()
    faza = Faza.get(Faza.detail == cor)
    fz = Faza.create(detail=index,
                    faza=pp.point.faza,
                    case=15,
                    weight=pp.point.assembly.weight,
                    in_work=faza.in_work,
                    in_object=faza.in_object,
                    weld=faza.weld,
                    kmd=faza.kmd,
                    set=faza.set,
                    preparation=faza.preparation,
                    shipment=faza.shipment,
                    assembly=faza.assembly,
                    paint=faza.paint,
                    mount=faza.mount,
                    packed=faza.packed
                    )

    paint = Detail.get(Detail.detail == cor,Detail.oper == 'paint')
    p = Detail.create(detail=index,basic=paint.basic,oper=paint.oper,to_work=0,norm=0,faza=fz)
    set = Detail.get(Detail.detail == cor,Detail.oper == 'set')
    d = Detail.create(detail=index,basic=set.basic,oper=set.oper,to_work=0,norm=0,faza=fz)
    index += 1


def LLL():
  norm = WeldNorm.select(WeldNorm.cathet).tuples()
  er = Weld.select().where(Weld.cathet.not_in(norm))
  for e in er:
    print(e,e.cathet)

  return
  points = Point.select().join(Drawing).where(Drawing.cas == 23,Point.faza == 1)
  for point in points:
    if PointPart.select().where(PointPart.point == point).first() == None:
      point.faza = 3
      point.save()
  return
  drawing = Drawing.get(Drawing.id == 3416)
  drawing.delete_instance(recursive=True)
  return
  import statistics
  drawing = Drawing.select(Drawing,fn.MAX(Drawing.area).alias('max'),fn.MIN(Drawing.area).alias('min'),fn.AVG(Drawing.area).alias('avg')).where(Drawing.cas == 15).first()
  print(drawing.min,drawing.avg,drawing.max)
  drawing = Drawing.select(Drawing.area).join(Part).tuples()
  aaa = []
  for d in drawing:
    aaa.append(int(d[0]))
  print(statistics.median(aaa))
  # print(statistics.median(drawing))

  # www = []
  # drawing = Drawing.select(Drawing,fn.SUM(Part.count).alias('aaa')).join(Part).where(Drawing.cas == 15).group_by(Drawing.id)
  # for d in drawing:
  #   www.append(d.aaa)
  #   if d.aaa > 100:
  #     print(d.assembly,d.aaa)
  # print(min(www))
  # print(max(www))
  # print(sum(www) / 2156)
  # print(statistics.median(www))
  return
  book = Workbook()
  sheet = book.active
  index = 1
  detail = 0
  pp = PointPart.select(PointPart,fn.COUNT(Part.id).alias('count')).join(Part).join(Drawing).where(Drawing.cas == 15,PointPart.weld == 0,PointPart.detail < 2493).group_by(PointPart.detail,Part.id)
  for p in pp:
    if p.detail != detail:
      sheet.cell(row=index,column=1).value = p.detail
      detail = p.detail
    sheet.cell(row=index,column=2).value = p.part.assembly.assembly
    sheet.cell(row=index,column=3).value = p.count
    index += 1
  book.save(f'noweld.xlsx')
    

  return
  faza = Faza.select().where(Faza.detail == 4945).first()
  faza.delete_instance(recursive=True)
  pp = PointPart.select().where(PointPart.detail == 4945)
  for p in pp:
    p.delete_instance(recursive=True)


  
  # task = TaskPart.select().join(Part).where(TaskPart.count != Part.count)
  # for t in task:
  #   print(t.part.assembly.assembly,t.part.number,t.count / t.part.assembly.count,t.part.count)
  # return
  # Faza.update({Faza.packed: 1}).where(Faza.packed == 3).execute()
  
  # return
  # Faza.update({Faza.faza: 39}).where(Faza.faza == 40).execute()
  # Point.update({Point.faza: 39}).where(Point.faza == 40).execute()
  # Task.update({Task.faza: 39}).where(Task.faza == 40).execute()
  
  return
  point = Point.select().join(Drawing).where(Drawing.cas == 15,Point.faza != 50)
  for i in point:
    pp = PointPart.select().join(Point).where(PointPart.point == i).first()
    if pp == None:
      # i.faza = 3
      # i.save()
      print(i.assembly.assembly,i.assembly.count,i.create_date,i.faza)



  # print(len(point))
  # faza = Faza.select().where(Faza.faza == 34)
  # print(len(faza))
  # print(len(pp))



  # return
  # tasks = Detail.select().where(Detail.end != None,Detail.oper != 'paint')
  # for task in tasks:
  #   if len(task.detailusers) == 0:
  #     print(task.detail,task.oper)

  #     data = []
  #     if task.worker_2 != None:
  #       data.append((task.id,task.worker_1,task.faza.weight / 2,task.norm / 2))
  #       data.append((task.id,task.worker_2,task.faza.weight / 2,task.norm / 2))
  #     else:
  #       data.append((task.id,task.worker_1,task.faza.weight,task.norm))
  #     print(data)
  #     with connection.atomic():
  #       DetailUser.insert_many(data,fields=[DetailUser.detail,DetailUser.worker,DetailUser.weight,DetailUser.norm]).execute()




  # return
  # detail = Faza.select(Faza.detail).where(Faza.faza != 32).tuples()
  # for d in detail:
  #   pp = PointPart.select().where(PointPart.detail == d).first()
  #   faza = Faza.select().where(Faza.detail == d).first()
  #   if pp.point.assembly.weight != faza.weight:
  #     faza.weight = pp.point.assembly.weight
  #     faza.save()
  #     print(d[0],pp.point.assembly.weight,faza.weight)

  # return
def ZAE():
  # pp = PointPart.select(PointPart.detail).group_by(PointPart.detail).tuples()
  # details = Faza.select().where(Faza.detail.not_in(pp))
  # for detail in details:
    # detail.delete_instance(recursive=True)
    # print(detail.detail)
  # return
  # point = Point.select().join(Drawing).where(Drawing.cas == 25,Point.faza == 1)
  # for i in point:
  #   pp = PointPart.select().join(Point).where(PointPart.point == i).first()
  #   if pp == None:
  #     i.faza = 2
  #     i.save()
  #     print(i.assembly.assembly,i.assembly.count,i.create_date,i.faza)
  # return
  # Faza.update({Faza.faza: 1}).where(Faza.faza == 2,Faza.case == 25).execute()
  # Point.update({Point.faza: 1}).where(Point.faza == 2,Point.assembly == 3797).execute()
  # Task.update({Task.faza: 1}).where(Task.faza == 2,Task.order == 25).execute()
  # return
  # drawing = Drawing.get(Drawing.id == 3440)
  # drawing.delete_instance(recursive=True)
  # return
  faza = 5
  case = 23
  for i in Faza.select().where(Faza.faza== faza,Faza.case == case):
    print(i)
    i.delete_instance(recursive=True)
  for i in Task.select().where(Task.faza == faza,Task.order == case):
    i.delete_instance(recursive=True)
  for i in Point.select().join(Drawing).where(Point.faza == faza,Drawing.cas == case):
    i.delete_instance(recursive=True)
  dr = Drawing.select().where(Drawing.cas == case)
  for d in dr:
    if d.count != len(d.points):
      print(d.assembly,d.count,len(d.points))
      if len(d.points) != 0:
        d.count = len(d.points)
        d.save()
      else:
        d.delete_instance(recursive=True)
  TaskPart.delete().where(TaskPart.part == None).execute()
  TaskPart.delete().where(TaskPart.task == None).execute()

  
def QWE():
  # otc = Otc.select(Otc,fn.COUNT(Otc.worker).alias('aaa')).where(Otc.end > datetime(2022,6,20,0,0,0),Otc.end < datetime(2022,7,2,23,0,0),Otc.oper == 'paint').group_by(Otc.worker)
  # for o in otc:
  #   print(o.worker.user.surname,o.aaa)
  
  # return
  wb = load_workbook('Склад.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Упаковка')
  excel = {}
  details = DetailPack.select().join(Packed).where(Packed.order == 15)
  for detail in details:
    i = False
    ppp = PointPart.select().join(Point).join(Drawing).where(PointPart.detail == detail.detail.detail,Drawing.cas == 15).group_by(PointPart.point)
    for pp in ppp:
      for y in range(3,285):
        for x in range(2,23):#23
          if pp.point.assembly.assembly == sheet.cell(row=y,column=x).value and sheet.cell(row=y,column=x).fill.start_color.rgb != '00FFFF00':
            print(pp.point.assembly.assembly)
            sheet.cell(row=y,column=x).fill = ft(start_color='FFFF00', end_color='0000FF', fill_type='solid')
            i = True
            break
        if i:
          break
  wb.save('Склад.xlsx')


def rrr():
  wb = load_workbook('Склад_clear.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Упаковка')

  details = Faza.select(Faza.detail).where(Faza.case == 15,Faza.packed == 1).tuples()
  ppp = PointPart.select(Drawing.assembly).join(Point).join(Drawing).where(PointPart.detail.in_(details)).group_by(PointPart.point).tuples()
  pp = []
  for p in ppp:
    pp.append(p[0])
  for y in range(3,285):
    for x in range(2,23):
      if sheet.cell(row=y,column=x).value != None:
        if sheet.cell(row=y,column=x).value in pp:
          sheet.cell(row=y,column=x).fill = ft(start_color='FF00FF', end_color='0000FF', fill_type='solid')
          pp.remove(sheet.cell(row=y,column=x).value)

  details = DetailPack.select(Faza.detail).join(Packed).join_from(DetailPack,Faza).where(Packed.order == 15).tuples()
  ppp = PointPart.select(Drawing.assembly).join(Point).join(Drawing).where(PointPart.detail.in_(details)).group_by(PointPart.point).tuples()
  pp = []
  for p in ppp:
    pp.append(p[0])
  for y in range(3,285):
    for x in range(2,23):
      if sheet.cell(row=y,column=x).value != None:
        if sheet.cell(row=y,column=x).value in pp:
          sheet.cell(row=y,column=x).fill = ft(start_color='FFFF00', end_color='0000FF', fill_type='solid')
          pp.remove(sheet.cell(row=y,column=x).value)


  details = DetailPack.select(Faza.detail).join(Packed).join_from(DetailPack,Faza).where(Packed.order == 15,Packed.shipment != None).tuples()
  ppp = PointPart.select(Drawing.assembly).join(Point).join(Drawing).where(PointPart.detail.in_(details)).group_by(PointPart.point).tuples()
  pp = []
  for p in ppp:
    pp.append(p[0])
  for y in range(3,285):
    for x in range(2,23):
      if sheet.cell(row=y,column=x).value != None:
        if sheet.cell(row=y,column=x).value in pp:
          sheet.cell(row=y,column=x).fill = ft(start_color='00FF00', end_color='0000FF', fill_type='solid')
          pp.remove(sheet.cell(row=y,column=x).value)


  wb.save('Склад.xlsx')


      # mark = sheet.cell(row=y,column=x).value




      # lmarks = []
      # for ex in excel.keys():
      #   lmarks.append(ex.lower())
      # if mark != None:
      #   if mark in excel.keys():
      #     dr = Drawing.get(Drawing.assembly == mark)
      #     excel[dr.assembly] += 1
      #   else:
      #     dr = Drawing.get(Drawing.assembly == mark)
      #     excel[dr.assembly] = 1
  
  # for drawing in Drawing.select(Drawing.assembly).where(Drawing.cas == 15).tuples():
  #   if drawing[0] in excel.keys() or drawing[0].startswith('МН'):
  #     pass
  #   else:
  #     print(drawing[0])

  # return
  # for e in excel:
  #   drawing = Drawing.select().where(Drawing.assembly == e,Drawing.cas == 15).first()
  #   try:
  #     if excel[e] != drawing.count:
  #       print(e,excel[e],drawing.count)
  #   except:
  #     print(e)

def PP():
  packs = Packed.select().where(Packed.ready == None,Packed.order == 35)
  for pac in packs:
    ready = PackList(pac)
    # pac.size = '0x0x0'
    pac.pack = 'Пакет'
    pac.ready = ready
    pac.save()
    dp = DetailPack.select(DetailPack.detail).where(DetailPack.pack == pac)
    for d in dp:
      d = d.detail
      d.packed = 3
      d.save()

def FOOD():
  details = Detail.select().where(Detail.detail >= 9045,Detail.detail <= 9045,Detail.oper == 'set')
  print(len(details))
  for detail in details:
    faza = detail.faza
    detail.to_work = 0
    detail.save()
    faza.preparation = 2
    faza.set = 0
    faza.save()

def NNN():
  for car in Shipment.select():
    detail = DetailPack.select(Faza.detail).join(Packed).join_from(DetailPack,Faza).where(Packed.shipment == car).tuples()
    Faza.update({Faza.shipment:3}).where(Faza.detail.in_(detail)).execute()
  return
  w2 = WeldNorm.select(WeldNorm.cathet).tuples()
  weld = Weld.select().where(Weld.cathet.not_in(w2))
  for w in w2:
    print(w)
  for we in weld:
    print(we.cathet)

def PL():
  pp = PointPart.select(PointPart.detail).join(Point).join(Drawing).where(Drawing.assembly == 'МУп-2').group_by(PointPart.detail).tuples()
  dp = DetailPack.select().join(Faza).where(Faza.detail.in_(pp))
  for d in dp:
    print(d.detail.detail,d.pack.number,d.pack.shipment)


def ASS():
  from decimal import Decimal
  er = []

  norms = Detail.select().where(Detail.norm != 0,Detail.oper == 'assembly')
  # norms = Detail.select().where(Detail.oper == 'assembly',Detail.detail == 2665)
  ind = len(norms)
  y = 2
  # book = Workbook()
  # sheet = book.active
  # sheet.cell(row=1,column=1).value = 'Наряд'
  # sheet.cell(row=1,column=2).value = 'Наименование конструкции'
  # sheet.cell(row=1,column=3).value = 'Наименование марки'
  # sheet.cell(row=1,column=4).value = 'Заказ'

  # S_min,S_max,t_min,t_max = (2,54,3,198)



  for norm in norms:
    print(ind)
    ind -= 1
    pp = PointPart.select(PointPart,fn.SUM(Part.count).alias('count'),fn.SUM(PointPart.bevel).alias('sum_bevel')).join(Part).where(PointPart.detail == norm.detail).first()

    # S = int(pp.count)

    # result = t_min-(((t_min-t_max)/100)*((100/(S_max-S_min))*(S-S_min)))

    # if pp.point.assembly.weight < 30:
    #   result = result * 0.7
    # elif pp.point.assembly.weight > 1000:
    #   result = result * 1.3

    result = 1

    if pp.sum_bevel > 0 and pp.point.name != 'Ограждение':
      result = float(1.3)

    print(result)

    an = AssemblyNorm.select().where(AssemblyNorm.name == pp.point.name,
                                     AssemblyNorm.mass_of <= pp.point.assembly.weight,
                                     AssemblyNorm.mass_to > pp.point.assembly.weight,
                                     AssemblyNorm.count_of <= pp.count,
                                     AssemblyNorm.count_to > pp.count).first()


    if an != None:
      # print(pp.point.name,pp.point.assembly.weight,pp.count,an.norm,an.norm /1000 * pp.point.assembly.weight,norm.detail)

      nnn = pp.point.assembly.weight * Decimal(result)
      norm.norm = an.norm / 1000 * nnn + 3
      norm.save()
      du = DetailUser.select().where(DetailUser.detail == norm)
      if len(du) != 0:
        DetailUser.update({DetailUser.norm: norm.norm / len(du)}).where(DetailUser.detail == norm).execute()


    else:
      er.append(('НЕТ В БАЗЕ',pp.point.name,pp.point.assembly.weight,pp.count,norm.detail))
      # sheet.cell(row=y,column=1).value = norm.detail
      # sheet.cell(row=y,column=2).value = pp.point.name
      # sheet.cell(row=y,column=3).value = pp.point.assembly.assembly
      # sheet.cell(row=y,column=4).value = pp.point.assembly.cas.cas
      y += 1
  for r in er:
    print(r)
  # book.save(f'Нет номенклатуры.xlsx')




def CORRECT():
  Point.update({Point.name: 'Связь вертикальная'}).where(Point.name == 'Связь верт.').execute()
  Point.update({Point.name: 'Ограждение'}).where(Point.name == 'Ограждения').execute()
  # AssemblyNorm.update({AssemblyNorm.norm:AssemblyNorm.norm * 1.2}).where(AssemblyNorm.name == 'Связь вертикальная').execute()
  # AssemblyNorm.update({AssemblyNorm.norm:AssemblyNorm.norm * 1.2}).where(AssemblyNorm.name == 'Связь').execute()
  # AssemblyNorm.update({AssemblyNorm.norm:AssemblyNorm.norm * 1.2}).where(AssemblyNorm.name == 'Связь горизонтальная').execute()
  # AssemblyNorm.update({AssemblyNorm.norm:AssemblyNorm.norm * 1.4}).where(AssemblyNorm.name == 'Площадка').execute()





def DELDEL(id):
  # drawings = Drawing.select().where(Drawing.cas == 30,Drawing.assembly.startswith('ФК3-1'))
  # for drawing in drawings:
  #   print(drawing.assembly)
  #   drawing.delete_instance(recursive=True)
  # fazas = Faza.select().where(Faza.detail >= 7355,Faza.detail <= 7360)
  # for faza in fazas:
  #   print(faza.detail)
  #   faza.delete_instance(recursive=True)
  # return
  order = Order.get(Order.cas == id)
  order.delete_instance(recursive=True)
  faza = Faza.select().where(Faza.case == None)
  for f in faza:
    f.delete_instance(recursive=True)
  TaskPart.delete().where(TaskPart.part == None).execute()
  TaskPart.delete().where(TaskPart.task == None).execute()
  task = TaskPart.select(Task.id).join(Task).tuples()
  tp = Task.select().where(Task.id.not_in(task))
  for t in tp:
    t.delete_instance()
  print(len(tp))



def TEST():
  # details = Faza.select(Faza.detail).where(Faza.case == 15).tuples()
  details = [1532,1703,1819,2120,2121,2156,2236,2418,2487,2679]
  avg = {'pp':0,'faza':0}
  for detail in details:
    pp = PointPart.select(PointPart,fn.SUM(Drawing.weight).alias('weight'),fn.SUM(PointPart.weld).alias('sum_weld')).join(Point).join(Drawing).where(PointPart.detail == detail).group_by(PointPart.detail).first()
    if pp.sum_weld == 0:
      x = 1
    else:
      x = pp.sum_weld
    
    faza = Faza.get(Faza.detail == detail)
    if pp.weight / x != faza.weight:
      print(pp.detail,pp.weight / x,faza.weight)
      avg['pp'] += pp.weight / x
      avg['faza'] += faza.weight
  print(avg)
  print(avg['faza'] - avg['pp'])


def PAINT():
  start = datetime(2022,7,1)
  end = datetime(2022,8,1)
  detail = Detail.select(Detail,fn.SUM(Faza.weight).alias('weight')).join(Faza).where(Detail.oper == 'set',Detail.worker_1 != None,Detail.end >= start,Detail.end < end).group_by(Detail.worker_1,Detail.worker_2).order_by(Detail.worker_1)
  detail2 = Detail.select(fn.SUM(Faza.weight).alias('weight')).join(Faza).where(Detail.oper == 'set',Detail.worker_1 != None,Detail.end >= start,Detail.end < end).scalar()
  book = Workbook()
  sheet = book.active
  y = 1 
  for d in detail:
    if d.worker_2 != None:
      user2 = d.worker_2.user.surname
    else:
      user2 = ''
    sheet.cell(row=y,column=1).value = d.worker_1.user.surname
    sheet.cell(row=y,column=2).value = user2
    sheet.cell(row=y,column=3).value = float(d.weight)
    print(d.worker_1.user.surname,user2,float(d.weight),d.detail)
    y += 1
  print(detail2)
  book.save(f'Комплектовщики.xlsx')
  
  return
  # pp = PointPart.select(PointPart.detail).tuples()
  # faza = Detail.select().where(Detail.detail.not_in(pp))
  # for f in faza:
  #   print(f.detail,f.case)
  # print(len(faza))

  # start = datetime(2022,7,1)
  # end = datetime(2022,8,1)

  # detail = Detail.select().where(Detail.oper == 'paint',Detail.worker_1 != None,Detail.end >= start,Detail.end < end).group_by(Detail.worker_1)
  # for d in detail:
  #   det = Detail.select(Detail.detail).where(Detail.worker_1 == d.worker_1,Detail.end >= start,Detail.end < end).tuples()
  #   pp = PointPart.select(fn.SUM(fn.DISTINCT(Drawing.weight))).join(Part).join(Drawing).where(PointPart.detail.in_(det),PointPart.weld == 1).scalar()
  #   print(d.worker_1.user.surname,d.worker_1.user.name,pp)

  # detail = Detail.select(Detail,fn.SUM(Faza.weight).alias('weight')).join(Faza).where(Detail.oper == 'paint',Detail.end >= start,Detail.end < end).group_by(Detail.worker_1)
  # for d in detail:
  #   print(d.worker_1.user.surname,d.worker_1.user.name,d.weight)
  # print('##################################')
  # detail = Detail.select(Detail,fn.SUM(Faza.area).alias('area')).join(Faza).where(Detail.oper == 'paint',Detail.end >= start,Detail.end < end).group_by(Detail.worker_1)
  # for d in detail:
  #   print(d.worker_1.user.surname,d.worker_1.user.name,d.area)


  data = []
  pointpart_paint = PointPart.select(PointPart,fn.SUM(Drawing.area).alias('area'),fn.SUM(PointPart.weld).alias('sum_weld')).join(Point).join(Drawing).group_by(PointPart.detail)
  for i in pointpart_paint:
    if i.sum_weld == 0:
      x = 1
    else:
      x = i.sum_weld
    data.append((i.detail,i.area / x))
  for d in data:
    Faza.update({Faza.area: d[1]}).where(Faza.detail == d[0]).execute()

def DEDR():
  # wb = load_workbook('Нет.xlsx',data_only=True)
  # sheet = wb.get_sheet_by_name('Sheet')
  # for i in range(2,116):
    # pp = PointPart.select().where(PointPart.detail == sheet.cell(row=i,column=1).value)
  # p = Point.select(Point.id).join(Drawing).join(Order).where(Order.cas == 2342).tuples()
  # for x in p:
  #   print(x.assembly.assembly)
  # Point.update({Point.name: 'Закладные'}).where(Point.id.in_(p)).execute()

  an = AssemblyNorm.select().where(AssemblyNorm.name == 'Ограждение')
  for a in an:
    AssemblyNorm.create(name='Арматурный каркас',mass_to=a.mass_to,mass_of=a.mass_of,count_to=a.count_to,count_of=a.count_of,norm=a.norm,complexity=1)


  # return
  # d = Faza.select().where(Faza.detail == 4983).first()
  # d.delete_instance(recursive=True)
  # return
  # fazas = Faza.select()
  # for faza in fazas:
  #   pp = PointPart.select().where(PointPart.detail == faza.detail).first()
  #   print(pp.point.assembly.cas.cas)

def WELDCOR():
  wb = load_workbook('Сварка заказ 23504.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Фаза3')
  for i in range(2,228):
    mark = (sheet.cell(row=i,column=1).value).replace(' ','')
    drawing = Drawing.get(Drawing.assembly == mark,Drawing.cas == 36)
    weld = Weld.select().where(Weld.assembly == drawing)
    print(len(weld),mark)
    Weld.delete().where(Weld.assembly == drawing).execute()
    if sheet.cell(row=i,column=2).value != None:
      Weld.create(assembly=drawing,cathet=4,length=sheet.cell(row=i,column=2).value)
    if sheet.cell(row=i,column=3).value != None:
      Weld.create(assembly=drawing,cathet=5,length=sheet.cell(row=i,column=3).value)
    if sheet.cell(row=i,column=4).value != None:
      Weld.create(assembly=drawing,cathet=6,length=sheet.cell(row=i,column=4).value)
    if sheet.cell(row=i,column=5).value != None:
      Weld.create(assembly=drawing,cathet=8,length=sheet.cell(row=i,column=5).value)
    if sheet.cell(row=i,column=6).value != None:
      Weld.create(assembly=drawing,cathet=9,length=sheet.cell(row=i,column=6).value)
    if sheet.cell(row=i,column=7).value != None:
      Weld.create(assembly=drawing,cathet=10,length=sheet.cell(row=i,column=7).value)
    if sheet.cell(row=i,column=8).value != None:
      Weld.create(assembly=drawing,cathet=12,length=sheet.cell(row=i,column=8).value)
    if sheet.cell(row=i,column=9).value != None:
      Weld.create(assembly=drawing,cathet=14,length=sheet.cell(row=i,column=9).value)
    if sheet.cell(row=i,column=10).value != None:
      Weld.create(assembly=drawing,cathet=16,length=sheet.cell(row=i,column=10).value)
    if sheet.cell(row=i,column=11).value != None:
      Weld.create(assembly=drawing,cathet=20,length=sheet.cell(row=i,column=11).value)

def WELDNORM():
  details = Detail.select().join(Faza).where(Detail.oper == 'weld')
  for detail in details:
    pp = PointPart.select().where(PointPart.detail == detail.detail).first()
    weld = Weld.select(fn.SUM(WeldNorm.norm * (Weld.length / 1000))).join(WeldNorm).where(Weld.assembly == pp.point.assembly).scalar()
    try:
      # if round(float(detail.norm),3) != round(float(weld),3):
        print(detail.detail)
        detail.norm = round(float(weld),3)
        detail.save()
        du = DetailUser.select().where(DetailUser.detail == detail)
        if len(du) != 0:
          DetailUser.update({DetailUser.norm: round(float(weld),3) / len(du)}).where(DetailUser.detail == detail).execute()
    except:
        print(detail.detail,detail.norm,weld)
  print('YES!!!!!')



def METALL():
  case = 2325
  faza = 39
  pp = PointPart.select(PointPart,Part,fn.SUM(Part.weight * Part.count).alias('weight')).join(Part).join(Drawing).join(Order).join_from(PointPart,Point).where(Order.cas == case,Point.faza == faza).group_by(Part.profile,Part.size,Part.mark)
  for p in pp:
    print(p.part.profile,p.part.size,p.part.mark,float(p.weight))



def NORMSAW():
  # TaskPart.delete().where(TaskPart.task == None).execute()
  # tt = TaskPart.select().where(TaskPart.part == None)
  # print(len(tt))
  profile = []
  tp = Task.select(Task,fn.SUM(TaskPart.count).alias('sum_count')).join(TaskPart).where(Task.oper.startswith('saw')).group_by(TaskPart.task)
  for t in tp:
    try:
      norm = SawNorm.select().where(SawNorm.profile == t.taskparts[0].part.profile,SawNorm.size == t.taskparts[0].part.size).first()
      (norm.norm_direct)
    except:
      if (t.taskparts[0].part.profile,t.taskparts[0].part.size) not in profile:
        profile.append((t.taskparts[0].part.profile,t.taskparts[0].part.size))
  book = Workbook()
  sheet = book.active
  index = 1
  for p in profile:
    sheet.cell(row=index,column=1).value = p[0]
    sheet.cell(row=index,column=2).value = p[1]
    index += 1
  book.save('Нет норм.xlsx')

def SHIP():
  book = Workbook()
  sheet = book.active
  fazas = Faza.select(Faza.detail).where(Faza.packed > 0, Faza.case == 33).tuples()
  pp = PointPart.select().where(PointPart.detail.in_(fazas)).group_by(PointPart.point)
  index = 1
  for p in pp:
    sheet.cell(row=index,column=1).value = index
    sheet.cell(row=index,column=2).value = p.point.assembly.assembly
    sheet.cell(row=index,column=3).value = float(p.point.assembly.weight)
    index += 1
  book.save('MARK.xlsx')
      
def KOZ(): # Отчет для Козликина
  book = Workbook()
  sheet = book.active
  index = 1
  start = datetime(2022,8,1)
  end = datetime(2022,8,31,23,59,59)
  noweld = Detail.select(Detail.detail).where(Detail.oper == 'weld').tuples()
  weld = Detail.select(Detail.detail).where(Detail.oper == 'weld',Detail.end > start,Detail.end < end).tuples()
  set = Detail.select(Detail.detail).where(Detail.oper == 'set',Detail.detail.not_in(noweld),Detail.end > start,Detail.end < end).tuples()
  details = Faza.select(Faza,fn.SUM(Faza.weight).alias('all_weight')).where((Faza.detail.in_(weld)) | (Faza.detail.in_(set))).group_by(Faza.case,Faza.faza)
  for detail in details:
    sheet.cell(row=index,column=1).value = detail.case.cas
    sheet.cell(row=index,column=2).value = detail.faza
    sheet.cell(row=index,column=3).value = float(detail.all_weight)
    index += 1
  book.save('REPORT.xlsx')

    
def WELDCOF():
  book = Workbook()
  sheet = book.active
  index = 1
  sheet.cell(row=index,column=1).value = 'Наряд'
  sheet.cell(row=index,column=2).value = 'Марка'
  sheet.cell(row=index,column=3).value = 'Чертеж'
  sheet.cell(row=index,column=4).value = 'Наименование'
  sheet.cell(row=index,column=5).value = 'Взял'
  sheet.cell(row=index,column=6).value = 'Сдал'
  sheet.cell(row=index,column=7).value = 'Дельта'
  sheet.cell(row=index,column=8).value = 'Дельта раб'
  sheet.cell(row=index,column=9).value = 'Норма'

  start = datetime(2022,9,1)
  end = datetime(2022,9,30,23,59,59)
  details = Detail.select(Detail,Faza).join(Faza).where((Detail.worker_1 == 2) | (Detail.worker_2 == 2),Detail.end > start, Detail.end < end).dicts()
  if len(details) != 0:
    for detail in details:
      pp = PointPart.select().where(PointPart.detail == detail['detail']).first()
      detail['mark'] = pp.point.assembly.assembly
      detail['draw'] = pp.point.draw
      detail['name'] = pp.point.name
      print(detail)
      index += 1
      sheet.cell(row=index,column=1).value = detail['detail']
      sheet.cell(row=index,column=2).value = pp.point.assembly.assembly
      sheet.cell(row=index,column=3).value = pp.point.draw
      sheet.cell(row=index,column=4).value = pp.point.name
      sheet.cell(row=index,column=5).value = detail['start']
      sheet.cell(row=index,column=6).value = detail['end']
      sheet.cell(row=index,column=7).value = detail['end'] - detail['start']
      if (detail['end'] - detail['start'] - timedelta(hours=16)) < timedelta(seconds=1):
        sheet.cell(row=index,column=8).value = detail['end'] - detail['start']
      else:
        sheet.cell(row=index,column=8).value = detail['end'] - detail['start'] - timedelta(hours=16)
      sheet.cell(row=index,column=9).value = f"{detail['norm'] // 60}:{int(detail['norm'] % 60)}"

  connection.close()
  book.save('WELD.xlsx')

def HOLECOF():
  # tas = Task.select().where(Task.oper == 'hole',Task.task == 2983)
  # for task in tasks:
  #   tps = TaskPart.select().where(TaskPart.task == task)
  #   for tp in tps:
  #     hole = Hole.select().where(Hole.part == tp.part)
  #     if len(hole) > 2:
  #       for h in hole:
  #         print(h.part)
  # for t in tas:
    # tasks = Task.select(Task,fn.SUM(Hole.count * TaskPart.count * HoleNorm.norm).alias('count')).join(TaskPart).join(Part).join(Hole).join(HoleNorm).where(Task.oper == 'hole',Task.id == t.id).group_by(Hole.diameter,Hole.depth)
    # tasks = Task.select(Task,fn.SUM(HoleNorm.norm).alias('count')).join(TaskPart).join(Part).join(Hole).join(HoleNorm).where(Task.oper == 'hole',Task.id == t.id).group_by(Hole.diameter,Hole.depth)
    # print(tasks[0],tasks[0].count)
    # break

  holes = Hole.select().where(Hole.norm == None)
  for hole in holes:
    if hole.part.profile == 'Лист':
      norm = HoleNorm.select().where(HoleNorm.depth_of <= hole.depth,
                                    HoleNorm.depth_to >= hole.depth,
                                    HoleNorm.diameter > hole.diameter,
                                    HoleNorm.count <= hole.count,
                                    HoleNorm.count_to > hole.count,
                                    HoleNorm.metal == 'Лист').first()
    else:
      norm = HoleNorm.select().where(HoleNorm.depth_of <= hole.depth,
                                    HoleNorm.depth_to >= hole.depth,
                                    HoleNorm.diameter > hole.diameter,
                                    HoleNorm.count <= hole.count,
                                    HoleNorm.count_to > hole.count,
                                    HoleNorm.lenght_of <= hole.part.length,
                                    HoleNorm.lenght_to > hole.part.length).first()
    hole.norm = norm
    print(hole.id)
  with connection.atomic():
    Hole.bulk_update(holes,fields=[Hole.norm])
    




def JKL():
  book = Workbook()
  sheet = book.active
  index = 1
  otc = Otc.select().where(Otc.fix == 1,Otc.oper == 'weld')
  for o in otc:
    sheet.cell(row=index,column=1).value = o.detail.detail
    sheet.cell(row=index,column=2).value = o.detail.case.cas
    sheet.cell(row=index,column=3).value = o.detail.packed
    print(o.detail.detail,o.detail.case.cas,o.detail.packed)
    index += 1
  print(len(otc))
  connection.close()
  book.save('Сварка.xlsx')

def PIN():
  pin = ('СПл-1','СПл-2','СПл-3','СПл-4','СПл-5','СПл-6','СПл-7')
  pp = PointPart.select(PointPart.detail).join(Part).join(Drawing).where(Drawing.assembly.in_(pin)).group_by(PointPart.detail).tuples()
  faza = Faza.select(Faza.id).where(Faza.detail.in_(pp)).tuples()
  # details = Detail.select().join(Faza).where(Detail.detail.in_(pp),Detail.oper == 'set',Detail.start == None)
  # for detail in details:
  #   detail.start = datetime.today()
  #   detail.end = datetime.today()
  #   detail.worker_1 = 85
  #   faza = detail.faza
  #   faza.set = 3
  #   faza.assembly = 3
  #   faza.weld = 3
  #   faza.paint = 3
  #   faza.packed = 1
  Detail.update({Detail.start:datetime.today(),Detail.end:datetime.today(),Detail.worker_1:92}).where(Detail.detail.in_(pp),Detail.start == None,Detail.oper == 'paint').execute()
  Detail.update({Detail.start:datetime.today(),Detail.end:datetime.today(),Detail.worker_1:85}).where(Detail.detail.in_(pp),Detail.start == None,Detail.oper == 'set').execute()
  Faza.update({Faza.set:3,Faza.assembly:3,Faza.weld:3,Faza.paint:3,Faza.packed:1}).where(Faza.detail.in_(pp)).execute()
  Otc.update({Otc.end:datetime.today(),Otc.worker:131}).where(Otc.detail.in_(faza),Otc.end == None).execute()

def HOLEC():
  HoleNorm.update({HoleNorm.norm:HoleNorm.norm * 4}).execute()
  HoleNorm.update({HoleNorm.norm:HoleNorm.norm * 1.3}).execute()
  # HoleNorm.update({HoleNorm.count:10,HoleNorm.count_to:30}).where(HoleNorm.count == 25,HoleNorm.count_to == 30).execute()
  # HoleNorm.update({HoleNorm.count:25,HoleNorm.count_to:30}).where(HoleNorm.count == 30,HoleNorm.count_to == 35).execute()
  # HoleNorm.update({HoleNorm.count:30,HoleNorm.count_to:1000}).where(HoleNorm.count == 35,HoleNorm.count_to == 1000).execute()
  # TaskPart.delete().where(TaskPart.part == None).execute()

def SAWCOF(profile,S,count,bevel=1):
  if profile == 'Двутавр':
    S_min,S_max,t_min,t_max = (10.32,187,2,25)
  if profile == 'Уголок':
    S_min,S_max,t_min,t_max = (1.47,78,0.7,15)
  if profile == 'Швеллер':
    S_min,S_max,t_min,t_max = (7.51,61,1.8,11)
  if profile == 'Труба круглая':
    S_min,S_max,t_min,t_max = (1.42,163,0.7,20)
  if profile == 'Труба профильная':
    S_min,S_max,t_min,t_max = (0.34,113,0.5,17)
  if profile == 'Круг':
    S_min,S_max,t_min,t_max = (0.79,707,0.8,40)

  S = float(S)
  result = t_min-(((t_min-t_max)/100)*((100/(S_max-S_min))*(S-S_min)))
  return result * int(count) * bevel

def TTT():
  book = Workbook()
  sheet = book.active
  index = 1
  # detail = Detail.select().where(Detail.oper == 'paint')


  ff = (15,23,25,27,29)
  faza = Faza.select().where(Faza.paint == 1,Faza.case.in_(ff))
  sheet.cell(row=index,column=1).value = 'Готов к покраске'
  for f in faza:
    print(f.detail)
    pp = list(PointPart.select(Drawing.assembly,Point.draw).join(Point).join(Drawing).where(PointPart.detail == f.detail).group_by(PointPart.point).tuples())
    index += 1
    sheet.cell(row=index,column=1).value = f'{f.detail} ({pp})'


  index = 1
  sheet.cell(row=index,column=2).value = 'В покраске'

  faza = Faza.select().where(Faza.paint == 2,Faza.case.in_(ff))
  for f in faza:
    index += 1
    pp = list(PointPart.select(Drawing.assembly,Point.draw).join(Point).join(Drawing).where(PointPart.detail == f.detail).group_by(PointPart.point).tuples())

    sheet.cell(row=index,column=2).value = f'{f.detail} ({pp})'
  book.save('reports/В малярке.xlsx')

  # Faza.update({Faza.packed:3}).where(Faza.shipment == 3).execute()
  # faza = Faza.select().where(Faza.shipment == 1,Faza.paint == 2).execute()
  # for f in faza:
    # f.packed = 3
    # f.save()
  #   print(f.detail)
  # print(len(faza))

  # dp = DetailPack.select(DetailPack.detail).tuples()
  # faza = Faza.select().where(Faza.paint == 1,Faza.id.in_(dp))
  # for f in faza:
  #   detail = Detail.select().where(Detail.detail == f.detail,Detail.oper == 'paint').first()
  #   detail.start = datetime.today()
  #   detail.end = datetime.today()
  #   detail.worker_1 = 92
  #   detail.save()
  #   f.paint = 3
  #   f.save()
  #   print(detail.detail)
  #   print(f.detail)

def JN():
  # wb = load_workbook('joint.xlsx',data_only=True)
  # sheet = wb.get_sheet_by_name('Лист1')
  # for y in range(1,32):
  #   JointNorm.create(
  #     profile=sheet.cell(row=y,column=1).value,
  #     size_of=sheet.cell(row=y,column=2).value,
  #     size_to=sheet.cell(row=y,column=3).value,
  #     norm=sheet.cell(row=y,column=4).value
  #   )
  
  lis = (8422,8423,8421,8420,8416,8411,8412,8408,8417,8415,8413,8410,8414,8418,8407,8424,8425,8426,8427,8429,8438,8461,8436,8440,8434,8439,8431,8428,8433,8430,8435,8462,8405,8468,8467,8441,8432,8409,8437,8472,8471)
  # print(len(lis))
  for i in range(8442,8461):
    # faza = Faza.select().where(Faza.detail == i)
    # DetailPack.create(detail=faza)
    print(i)
    detail = Detail.select().where(Detail.detail == i,Detail.oper == 'paint').first()
    faza = detail.faza
    detail.start = datetime.today()
    detail.end = datetime.today()
    detail.worker_1 = 92
    faza.paint = 3
    faza.packed = 1
    detail.save()
    faza.save()

def KK():
  du = DetailUser.select(DetailUser,fn.COUNT(DetailUser.detail).alias('count')).group_by(DetailUser.detail,DetailUser.worker)
  for d in du:
    if d.count > 1:
      DetailUser.delete().where(DetailUser.id == d.id).execute()
      print(d.detail,d.norm,d.detail.norm)

def JJ():
  # wb = load_workbook('VVL3.xlsx',data_only=True)
  # sheet = wb.get_sheet_by_name('Сборка (ЕНиР)')
  # for i in range(18,63):
  #   assembly = AssemblyNorm.get(
  #     AssemblyNorm.name == 'Колонна',
  #     AssemblyNorm.mass_of == sheet.cell(row=i,column=11).value,
  #     AssemblyNorm.mass_to == sheet.cell(row=i,column=12).value,
  #     AssemblyNorm.count_of == sheet.cell(row=i,column=13).value,
  #     AssemblyNorm.count_to == sheet.cell(row=i,column=14).value,
  #   )
  #   assembly.norm = sheet.cell(row=i,column=16).value
  #   assembly.save()
  # AssemblyNorm.delete().where(AssemblyNorm.name == 'Стойка').execute()
  # points = Point.select().join(Drawing).where(Drawing.cas == 34,Point.name == 'Стойка фахверка')
  # for point in points:
  #   print(point)
  #   point.name = 'Стойка'
  #   point.save()
  wb = load_workbook('Очередность 2345.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Упаковка')
  lis = []
  for x in range(1,12):
    for y in range(4,99):
      p = sheet.cell(row=y,column=x).value
      if p != None:
        pp = PointPart.select(PointPart.detail).join(Point).join(Drawing).where(Drawing.assembly == p,PointPart.detail.not_in(lis)).group_by(PointPart.detail).tuples()
        if len(pp) != 0:
          faza = Faza.select().where(Faza.detail.in_(pp),Faza.paint == 1).first()
          if faza != None:
            lis.append(faza.detail)
            print(faza.detail)
            sheet.cell(row=y,column=x).fill = ft(start_color='00FF00', end_color='0000FF', fill_type='solid')
            sheet.cell(row=y,column=x).value = f'{p}({faza.detail})'
          else:
            faza = Faza.select().where(Faza.detail.in_(pp),Faza.weld == 2).first()
            if faza != None:
              lis.append(faza.detail)
              print(faza.detail)
              sheet.cell(row=y,column=x).fill = ft(start_color='0000FF', end_color='0000FF', fill_type='solid')
              sheet.cell(row=y,column=x).value = f'{p}({faza.detail})'
            else:
              faza = Faza.select().where(Faza.detail.in_(pp),Faza.weld == 1).first()
              if faza != None:
                lis.append(faza.detail)
                print(faza.detail)
                sheet.cell(row=y,column=x).fill = ft(start_color='FF0000', end_color='0000FF', fill_type='solid')
                sheet.cell(row=y,column=x).value = f'{p}({faza.detail})'
              else:
                faza = Faza.select().where(Faza.detail.in_(pp),Faza.assembly == 2).first()
                if faza != None:
                  lis.append(faza.detail)
                  print(faza.detail)
                  sheet.cell(row=y,column=x).fill = ft(start_color='FFFF00', end_color='0000FF', fill_type='solid')
                  sheet.cell(row=y,column=x).value = f'{p}({faza.detail})'
                else:
                  faza = Faza.select().where(Faza.detail.in_(pp),Faza.assembly == 1).first()
                  if faza != None:
                    lis.append(faza.detail)
                    print(faza.detail)
                    sheet.cell(row=y,column=x).fill = ft(start_color='FF00FF', end_color='0000FF', fill_type='solid')
                    sheet.cell(row=y,column=x).value = f'{p}({faza.detail})'
                  else:
                    faza = Faza.select().where(Faza.detail.in_(pp),Faza.set == 2).first()
                    if faza != None:
                      lis.append(faza.detail)
                      print(faza.detail)
                      sheet.cell(row=y,column=x).fill = ft(start_color='00FFFF', end_color='0000FF', fill_type='solid')
                      sheet.cell(row=y,column=x).value = f'{p}({faza.detail})'
                    else:
                      faza = Faza.select().where(Faza.detail.in_(pp),Faza.set == 1).first()
                      if faza != None:
                        lis.append(faza.detail)
                        print(faza.detail)
                        sheet.cell(row=y,column=x).fill = ft(start_color='004444', end_color='0000FF', fill_type='solid')
                        sheet.cell(row=y,column=x).value = f'{p}({faza.detail})'
                      else:
                        faza = Faza.select().where(Faza.detail.in_(pp),Faza.preparation == 2).first()
                        if faza != None:
                          lis.append(faza.detail)
                          print(faza.detail)
                          sheet.cell(row=y,column=x).fill = ft(start_color='FFFFFF', end_color='0000FF', fill_type='solid')
                          sheet.cell(row=y,column=x).value = f'{p}({faza.detail})'
  wb.save('Очередность.xlsx')


def VVV():
  # fazas = Faza.select().where(Faza.shipment != 0,Faza.packed != 3)
  # for faza in fazas:
    # print(faza.detail)
  # print(len(fazas))
  # Faza.update({Faza.packed:3}).where(Faza.shipment != 0,Faza.packed != 3).execute()
  # bolts = Nut.select(Nut.profile,fn.SUM(Nut.count * Drawing.count).alias('count_all')).join(Drawing).join(Order).where(Order.cas == '23504').group_by(Nut.profile)
  # all = 0
  # for bolt in bolts:
    # all += bolt.count_all
    # print(bolt.profile,bolt.count_all)
  # print(all)
  # Drawing.update({Drawing.paint:'1'}).execute()
  fazas = Faza.select(Faza.case,fn.MIN(Faza.shipment).alias('end')).group_by(Faza.case)
  for faza in fazas:
    if faza.end == 3 and faza.case.status != 'Готов':
      order = faza.case
      order.status = 'Готов'
      order.color = 'white'
      order.save()
      print(faza.case.cas,faza.end)

def ZXC():
  # marks = ('СПл-1','СПл-2','СПл-3','СПл-4','СПл-5','СПл-6','СПл-7')
  # pp = PointPart.select(PointPart.detail).join(Point).join(Drawing).where(Drawing.assembly.in_(marks)).group_by(PointPart.detail).tuples()
  # pack = Packed.select().join(DetailPack).join(Faza).where(Faza.detail.in_(pp)).group_by(Packed.id)
  # for p in pack:
    # p.ready = None
    # p.save()
    # print(p.number)
  # drawing = PointPart.select(PointPart.detail).join(Point).join(Drawing).where(Drawing.assembly.startswith('МОг-'),Drawing.cas == 15).group_by(PointPart.detail).tuples()
  # fazas = Faza.select(Faza.id).where(Faza.detail.in_(drawing)).tuples()
  # print(len(fazas))
  # Faza.update({Faza.packed: 1}).where(Faza.detail.in_(drawing)).execute()
  # dps = DetailPack.select(DetailPack.pack).join(Faza).where(Faza.detail.in_(drawing)).group_by(DetailPack.pack).tuples()
  # print(len(dps))
  # packs = Packed.select().where(Packed.id.in_(dps))
  # for pack in packs:
  pack = Drawing.get(Drawing.id == 6200)
  print(pack)
  pack.delete_instance(recursive=True)

def CXZ():
  wb = load_workbook('металл.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Лист1')
  for y in range(1,1934):
    sheet.cell(row=y,column=2).value = (sheet.cell(row=y,column=2).value).strip()
    if sheet.cell(row=y,column=2).value == 'Труба квдратная':
      sheet.cell(row=y,column=2).value = 'Труба квадратная'
      text = sheet.cell(row=y,column=3).value 
      sheet.cell(row=y,column=3).value = text[text.find('*')+1:]
    if sheet.cell(row=y,column=2).value == 'Уголок равнополочный':
      text = sheet.cell(row=y,column=3).value 
      sheet.cell(row=y,column=3).value = text[text.find('*')+1:]
    sheet.cell(row=y,column=3).value = str(sheet.cell(row=y,column=3).value).replace('*','x')
    if sheet.cell(row=y,column=4).value == 'ст0-ст5':
      sheet.cell(row=y,column=4).value = 'ст0,ст1,ст2,ст3,ст4,ст5'
    if (sheet.cell(row=y,column=2).value).find('Арматура') != -1:
      sheet.cell(row=y,column=8).value = 'Арматура'
    elif (sheet.cell(row=y,column=2).value).find('двутавр') != -1:
      sheet.cell(row=y,column=8).value = 'Двутавр'
    elif (sheet.cell(row=y,column=2).value).find('Круг') != -1:
      sheet.cell(row=y,column=8).value = 'Круг'
    elif (sheet.cell(row=y,column=2).value).find('Уголок') != -1:
      sheet.cell(row=y,column=8).value = 'Уголок'
    elif (sheet.cell(row=y,column=2).value).find('Труба квадратная') != -1:
      sheet.cell(row=y,column=8).value = 'Труба профильная'
    elif (sheet.cell(row=y,column=2).value).find('Труба прямоугольная') != -1:
      sheet.cell(row=y,column=8).value = 'Труба профильная'
    elif (sheet.cell(row=y,column=2).value).find('Труба') != -1:
      sheet.cell(row=y,column=8).value = 'Труба круглая'
    elif (sheet.cell(row=y,column=2).value).find('Лист просечно-вытяжной') != -1:
      sheet.cell(row=y,column=8).value = 'Лист ПВ'
    elif (sheet.cell(row=y,column=2).value).find('рифление') != -1:
      sheet.cell(row=y,column=8).value = 'Лист РИФ'
    elif (sheet.cell(row=y,column=2).value).find('Швеллер гнутый') != -1:
      sheet.cell(row=y,column=8).value = 'Швеллер гнутый'
    elif (sheet.cell(row=y,column=2).value).find('Швеллер') != -1:
      sheet.cell(row=y,column=8).value = 'Швеллер'
    elif (sheet.cell(row=y,column=2).value).find('Лист') != -1:
      sheet.cell(row=y,column=8).value = 'Лист'
    elif (sheet.cell(row=y,column=2).value).find('Шестигранник') != -1:
      sheet.cell(row=y,column=8).value = 'Шестигранник'
    elif (sheet.cell(row=y,column=2).value).find('Квадрат') != -1:
      sheet.cell(row=y,column=8).value = 'Квадрат'
    elif (sheet.cell(row=y,column=2).value).find('Сетка') != -1:
      sheet.cell(row=y,column=8).value = 'Сетка'
      sheet.cell(row=y,column=7).value = 0
  wb.save('steel.xlsx')


def Catalog():
  wb = load_workbook('steel.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Лист1')
  list = []
  for y in range(1,1934):
    width = False
    if sheet.cell(row=y,column=6).value == 'м2':
      width = True
    list.append((
      sheet.cell(row=y,column=8).value,
      sheet.cell(row=y,column=2).value,
      sheet.cell(row=y,column=3).value,
      sheet.cell(row=y,column=4).value,
      sheet.cell(row=y,column=5).value,
      width,
      float(sheet.cell(row=y,column=7).value),
      ))
  with connection.atomic():
    CatalogSteel.insert_many(list, fields=[CatalogSteel.name,CatalogSteel.full_name,CatalogSteel.size,CatalogSteel.mark,CatalogSteel.gost,CatalogSteel.width,CatalogSteel.weight]).execute()

def DelMark():
  details = (13773,13850,13909,13859,13861,13858,13860,13882,13887,13814,13813,13815,13816,13817,13811,13820,13818,13819,13812,13954,13953,13951,13821,13881,13810,13880,13882,13883,13884,13885,13955)
  for d in details:
    try:
      pp = PointPart.select().where(PointPart.detail == d).group_by(PointPart.point)
      faza = Faza.get(Faza.detail == d)
      for p in pp:
        point = p.point
        point.delete_instance(recursive=True)
      faza.delete_instance(recursive=True)
    except:
      print(d,'NO')
  drawing = Drawing.select().where(Drawing.assembly == 'ДСф-22').first()
  if drawing != None:
    drawing.delete_instance(recursive=True)

def JKL():
  # fazas = Faza.select().where(Faza.case == 35)
  # for faza in fazas:
    # c = 0 
    # pp = PointPart.select().join(Point).join(Drawing).where(PointPart.detail == faza.detail).group_by(PointPart.point)
    # for p in pp:
      # c += p.point.assembly.weight
    # faza.weight = c
  # with connection.atomic():
    # Faza.bulk_update(fazas, fields=[Faza.weight])

  # return
  for order in Order.select():
    weight_all = Point.select(Point.faza,fn.SUM(Drawing.weight).alias('weight')).join(Drawing).join(Order).where(Order.id == order.id).first()
    order.weight = weight_all.weight
    order.save()
    print(order.cas,weight_all.weight)

def LKJ():
  details = Detail.select(Detail.detail).join(Faza).where(Faza.case == 51,Detail.oper == 'paint').tuples()
  details = Detail.select(Detail.detail,Faza.id).join(Faza).where(Faza.case == 51,Detail.oper == 'set',Detail.detail.not_in(details)).tuples()
  for i in details:
    print(i)
    Detail.create(detail=i[0],basic='paint',oper='paint',to_work=0,norm=0,faza=i[1])

def exl():
  book = Workbook()
  sheet = book.active
  points = Point.select(Point).join(Drawing).where(Drawing.cas == 15).group_by(Drawing.assembly)
  r = 1
  for i in points:
    print(i.assembly.assembly,i.draw)
    sheet.cell(row=r,column=1).value = r
    sheet.cell(row=r,column=2).value = i.assembly.assembly
    sheet.cell(row=r,column=3).value = i.name
    sheet.cell(row=r,column=4).value = i.draw
    sheet.cell(row=r,column=5).value = i.assembly.count
    sheet.cell(row=r,column=6).value = i.assembly.weight
    sheet.cell(row=r,column=7).value = i.assembly.weight * i.assembly.count
    r += 1
  book.save('list.xlsx')


def MMM():
  book = Workbook()
  sheet = book.active
  case = '2325'
  pp = PointPart.select(Part.profile,Part.size,Part.mark,fn.SUM(Part.weight * Part.count).alias('weight')).join(Part).join(Drawing).join(Order).join_from(PointPart,Point).where(Order.cas.startswith(case)).group_by(Part.profile,Part.size,Part.mark).dicts()
  r = 1
  for i in pp:
    sheet.cell(row=r,column=1).value = i['profile']
    sheet.cell(row=r,column=2).value = i['size']
    sheet.cell(row=r,column=3).value = i['mark']
    sheet.cell(row=r,column=4).value = i['weight']
    r += 1
  book.save('555.xlsx')
  # t = (2051,2093,2096,2107)
  return
  # Packed.update(ready=None).where(Packed.number.in_(t)).execute()

  # return
  # detail = PointPart.select(PointPart.detail).join(Point).join(Drawing).where(Drawing.assembly == 'СПл-9').group_by(PointPart.detail).limit(72).tuples()
  detail = PointPart.select(PointPart.detail).join(Point).join(Drawing).where(Drawing.assembly == 'СШ-1').group_by(PointPart.detail).tuples()
  t = list(detail)


  details = DetailPack.select(DetailPack,fn.COUNT(DetailPack.id).alias('count')).join(Faza).where(Faza.detail.in_(t)).group_by(DetailPack.pack)
  # details = DetailPack.select().join(Faza).where(Faza.detail.in_(t))#.group_by(DetailPack.pack)
  for i in details:
    # i.pack = 2220
    # i.save()
    print(i.pack.number,i.count)
    Packed.update(ready=None).where(Packed.number == i.pack.number).execute()
    # print(i.pack.number)

def Fask():
  # pp = PointPart.select(PointPart.detail).join(Point).join(Drawing).where(Drawing.assembly == 'МС-2').group_by(PointPart.detail).tuples()
  # for i in pp:
    # print(i)
    # i.delete_instance(recursive=True)
  faza = Faza.select(Faza.detail).join(Order).where(Order.cas == '23256',Faza.set == 1).tuples()
  pp = PointPart.select(Point).join(Point).join(Drawing).where(PointPart.detail.in_(faza)).group_by(PointPart.point)
  for i in pp:
    print(i.point)
    point = i.point
    point.delete_instance(recursive=True)
  faza = Faza.select().join(Order).where(Order.cas == '23256',Faza.set == 1)
  for i in faza:
    i.delete_instance(recursive=True)


  return
  pp = PointPart.select().join(Part).join(Drawing).join(Order).where(Order.cas == '1359',Part.sn.in_(('18','19','20','21','22')),Drawing.assembly == 'БП2')
  for p in pp:
    print(p.detail)
    p.joint = 1
    p.save()

def Deltask():
  ship = Shipment.select().join(Packed).join(DetailPack).join(Faza).where(Shipment.number == 156).group_by(Faza.detail)
  t = []
  for packs in ship:
    for dp in packs.packeds:
      if dp.id not in t:
        t.append(dp.id)
        print(dp.shipment)
        dp.shipment = None
        dp.save()
  return
  CatalogSteel.update(gost='30245-2012').where(CatalogSteel.gost=='30245-2003').execute()
  return
  fazas = Faza.select().where(Faza.paint == 1)
  for faza in fazas:
    # Detail.delete().where(Detail.oper == 'paint',Detail.detail == faza.detail).execute()
    det = Detail.select().where(Detail.detail == faza.detail,Detail.oper == 'paint').first()
    if det == None:
      print(faza.detail)
      # faza.paint = 3
      # faza.save()
      # Otc.create(detail=faza,oper='paint')
    # otc = Otc.select().where(Otc.detail == faza,Otc.oper == 'paint')
    # for i in otc:
      # print(i.detail.detail,i.oper)
      # i.delete_instance()
  return
  points = Point.select(Drawing.id).join(Drawing).group_by(Drawing.id).tuples()
  drawings = Drawing.select().where(Drawing.id.not_in(points))
  for i in drawings:
    print(i.cas,i.assembly)
    # i.delete_instance(recursive=True)
  orders = Drawing.select(Drawing,fn.COUNT(Point.id).alias('aaa')).join(Point).group_by(Drawing.id)
  for i in orders:
    if i.aaa == 0:
      print(i.id,i.assembly)
  # points = Point.select().join(Drawing).join(Order).where(Order.cas == '1510',Point.faza == 4)
  # for point in points:
    # point.delete_instance(recursive=True)
  # fazas = Faza.select().join(Order).where(Order.cas == '1510',Faza.faza == 4)
  # for point in fazas:
    # point.delete_instance(recursive=True)
  # for d in dd:

    # if d.area > 0:
      # d.upload = 'tekla'
    # else:
      # d.upload = 'manual'
    # d.save()
  pass
  # faza = Faza.select(Faza.detail).where(Faza.case == 67).tuples()
  # notdetail = Detail.select(Detail.detail).where(Detail.detail.in_(faza),Detail.oper=='paint').tuples()
  # faza = Faza.select(Faza.detail).where(Faza.detail.in_(faza),Faza.detail.not_in(notdetail)).tuples()
  # for f in faza:
    # Detail.create(detail=f[0],basic='paint',oper='paint',to_work=0,norm=0,faza=18756)

def SUD():
  pack = Packed.select().where(Packed.order == 35)
  for p in pack:
    print(p.number)
    p.ready = None
    p.save()
  return
  faza = Faza.select(Faza.detail).where(Faza.case == 33,Faza.shipment != 3).tuples()
  pp = PointPart.select().join(Point).join(Drawing).where(PointPart.detail.in_(faza)).group_by(Point.id)
  ppp = []
  for p in pp:
    ppp.append(p.point.id)
  point = Point.select(Point,fn.COUNT(Drawing.assembly).alias('count_as')).join(Drawing).where(Point.id.in_(ppp)).group_by(Drawing.assembly)
  book = Workbook()
  sheet = book.active
  r = 1
  for f in point:
    print(f.assembly.assembly)
    sheet.cell(row=r,column=1).value = f.assembly.assembly
    sheet.cell(row=r,column=2).value = f.count_as
    sheet.cell(row=r,column=3).value = f.assembly.weight
    sheet.cell(row=r,column=4).value = f.assembly.weight * f.count_as
    r += 1
  book.save('sud2.xlsx')

def DP():
  pp = Packed.select().where(Packed.order == 27)
  i = len(pp)
  for p in pp:
    detail = DetailPack.select(Faza.detail).join(Faza).where(DetailPack.pack == p).tuples()
    Faza.update({Faza.packed:1}).where(Faza.detail.in_(detail)).execute()
    DetailPack.delete().where(DetailPack.pack == p).execute()
    p.delete_instance(recursive=True)
    print(i)
    i -= 1

def ggg():
  import xmltodict
  import json
  with open('contact.xml',encoding="utf-8") as file:
    doc = xmltodict.parse(file.read())

  j = json.dumps(doc,ensure_ascii=False)
  print(j)
  return
  for i in doc['PhoneBook']['DirectoryEntry']:
    name = i['Name'].split(' ')[0]
    print(name,'-',i['Telephone'])
    user = User.select().where(User.surname == name).first()
    if user != None:
      PhoneBook.create(surname=user.surname,name=user.name,patronymic=user.patronymic,phone=i['Telephone'],direction='in')
    else:
      PhoneBook.create(company=name,phone=i['Telephone'],direction='in')

def ggg2():
  import xmltodict
  import pysftp
  xmls = ('in','out','spec')
  for text in xmls:
    phone_in = {'PhoneBook':{"DirectoryEntry":[]}}
    phones = PhoneBook.select().where(PhoneBook.direction == text).order_by(
        fn.CONCAT(
            (Case(None,[(PhoneBook.company == None,'')],PhoneBook.company)),
            (Case(None,[(PhoneBook.surname == None,'')],PhoneBook.surname))
          )
      )
    for phone in phones:
      st = ''
      if phone.company:
        st += phone.company
      if phone.surname:
        st += ' '
        st += phone.surname +' '+ phone.name[0] +'.'+phone.patronymic[0]
      st = st.strip()
      p = {"Name":st,"Telephone":phone.phone,"Mobile":None,"Other":None,"Ring":"Default","Group":None}
      phone_in["PhoneBook"]["DirectoryEntry"].append(p)
    xml = (xmltodict.unparse(phone_in,pretty=True))
    with open(f'{text}.xml','w') as f:
      f.write(xml)
    with pysftp.Connection('192.168.0.51',username='admin',password='15296MV6') as sftp:
      with sftp.cd('123'):
        sftp.put(f'{text}.xml')
