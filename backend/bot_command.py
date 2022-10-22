from email import message_from_binary_file
from openpyxl import Workbook, load_workbook
from models import *
from qr_list import QR_pdf
from send_bot import AllReport, Report, Viktor, ViktorDocs
from scaner import Scaner, Scaner2
from datetime import datetime, timedelta
from peewee import fn


async def Aut(message):
  return User.select().where(User.telegram == message.from_user.id).first() == None

def Group(message,group):
  return Worker.select(User.telegram).join(User).where(User.telegram == message.chat.id,Worker.oper.in_(group)).first()

async def Marka(text):
  pp = PointPart.select(PointPart.detail).join(Point).join(Drawing).join(Order).where(Drawing.assembly == text[1],Order.cas == text[2]).group_by(PointPart.detail).tuples()
  faza = Faza.select().where(Faza.detail.in_(pp))
  result = ''
  status = ''
  for stage in faza:
    if stage.set == 0:
      status = 'Заготовка'
    elif stage.set == 1:
      status = 'Готов к комплектации'
    elif stage.set == 2:
      status = 'В комплектации'
    elif stage.assembly == 1:
      status = 'Готов к сборке'
    elif stage.assembly == 2:
      status = 'В сборке'
    elif stage.weld == 1:
      status = 'Готов к сварке'
    elif stage.weld == 2:
      status = 'В сварке'
    elif stage.paint == 1:
      status = 'Готов к покраске'
    elif stage.paint == 2:
      status = 'В покраске'
    elif stage.packed == 1:
      status = 'Готов к упаковке'
    elif stage.packed == 3 and stage.shipment == 0:
      status = 'Упакован'
    elif stage.packed == 3 and stage.shipment == 3:
      status = 'Отгружен'
    else:
      status = '?'
    result += f'{stage.detail} {status}\n'
  return result

async def Registration(message,text):
  reg = Reg.select().where(Reg.telegram == message.chat.id).first()
  if reg:
    return 'Заявка уже подана'
  user = User.select().where(User.telegram == message.chat.id).first()
  if user:
    return 'Вы уже зарегистрированны'
  reg = Reg.create(surname=text[1],name=text[2],patronymic=text[3],telegram=message.chat.id,username=message.chat.username)
  user = User.select().where(User.surname == text[1],User.telegram.in_([0,None])).first()
  if user:
    reg.user = user
    reg.save()
    if user.surname == message.chat.last_name and user.name == message.chat.first_name[0]:
      user.name = text[2]
      user.patronymic = text[3]
      user.telegram = message.chat.id
      user.username = message.chat.username
      user.save()
      Reg.delete().where(Reg.id == reg.id).execute()
      return f'регистрация прошла успешно'
    else:
      await Viktor(f'регистрация {reg.id}\n{text[1]}\n{text[2]}\n{text[3]}\nВ базе {reg.user}')
  else:
    await Viktor(f'регистрация {reg.id}\n{text[1]}\n{text[2]}\n{text[3]}\nВ базе {reg.user}')
  result = f'Заявка на регистрацию отправлена, ожидайте подтверждения'
  return result

async def RegMan(text):
  reg = Reg.select().where(Reg.id == text[1]).first()
  if reg and reg.user:
    user = User.get(User.id == reg.user.id)
    user.name = reg.name
    user.patronymic = reg.patronymic
    user.telegram = reg.telegram
    user.username = reg.username
    user.save()
    Reg.delete().where(Reg.id == reg.id).execute()
    await Report(user.telegram,'регистрация прошла успешно')
    return 'регистрация прошла успешно'
  elif reg and reg.user == None:
    user = User.create(surname=reg.surname,name=reg.name,patronymic=reg.patronymic,telegram=reg.telegram,username=reg.username)
    Reg.delete().where(Reg.id == reg.id).execute()
    await Report(user.telegram,'регистрация прошла успешно')
    return 'регистрация прошла успешно'
  else:
    return 'Нет такого запроса'


async def Photo(message):
  
  if message.photo:
    file_name = f'media/scaner/{message.chat.id}.jpg'
    await message.photo[-1].download(file_name)
    qr = Scaner(file_name)
  elif message.video:
    file_name = f'media/scaner/{message.chat.id}.xlsx'
    await message.video.download(file_name)
    qr = Scaner2(file_name)
  text = qr.split(' ')
  print(text)
  return text

async def Workers(message,text):
  date = datetime.today()
  if text[0] == 'Run' or text[0] == 'A':
    detail = Detail.select().where(Detail.detail == text[1],Detail.to_work == 1,Detail.start == None).first()

    if detail == None:
      detail = Detail.select().where(Detail.detail == text[1],Detail.to_work == 1,Detail.start != None,Detail.end == None).first()

      if detail == None:
        return 'Все операции по данному наряду закончены'
      elif detail.worker_1.user.telegram != message.chat.id and detail.worker_2 == None and (date - detail.start > timedelta(minutes=10)):
        return 'Заказ в работе более 10 минут! Вы не можете присоединиться'
      elif detail.worker_1.user.telegram != message.chat.id and detail.worker_2 != None and detail.worker_2.user.telegram != message.chat.id:
        return 'Заказ уже в работе'
      elif (detail.worker_1.user.telegram == message.chat.id) or (detail.worker_2 != None and detail.worker_2.user.telegram == message.chat.id):
        if date - detail.start < timedelta(minutes=1):
          return 'Минимальное время 30 минут'
        detail.end = date
        detail.save()
        faza = detail.faza
        if detail.oper == 'set' and Detail.select().where(Detail.oper == 'assembly',Detail.detail == detail.detail).first() != None:
          Detail.update({Detail.to_work: 1}).where(Detail.oper == 'assembly',Detail.detail == detail.detail).execute()
          faza.set = 3
          faza.assembly = 1
          faza.save()
          DetailUserAdd(detail)
          return 'Комплектовка завершина'
        elif detail.oper == 'set' and Detail.select().where(Detail.oper == 'assembly',Detail.detail == detail.detail).first() == None:
          Detail.update({Detail.to_work: 1}).where(Detail.oper == 'paint',Detail.detail == detail.detail).execute()
          faza.set = 3
          faza.assembly = 3
          faza.weld = 3
          faza.paint = 1
          faza.save()
          DetailUserAdd(detail)
          return 'Комплектовка завершина'
        elif detail.oper == 'assembly':
          Detail.update({Detail.to_work: 1}).where(Detail.oper == 'weld',Detail.detail == detail.detail).execute()
          faza.assembly = 3
          faza.weld = 1
          faza.save()
          DetailUserAdd(detail)
          return 'Сборка завершина'

        elif detail.oper == 'weld':
          # Detail.update({Detail.to_work: 1}).where(Detail.oper == 'paint',Detail.detail == detail.detail).execute()
          otc = Otc.select().where(Otc.detail == faza,Otc.oper == 'weld').first()
          if otc == None:
            otc = Otc.create(detail=faza,start=datetime.today(),oper='weld')
            faza.weld = 3
            faza.save()
            otc = Worker.select(User.telegram).join(User).where(Worker.oper.in_(['otc'])).tuples()
            await AllReport(otc,f'{faza.detail} {detail.worker_1.user.surname} Наряд на проверку сварки')
            DetailUserAdd(detail)
          else:
            otc.fix = 0
            otc.error += 1
            otc.save()
          return 'Сварка завершина'

        elif detail.oper == 'paint':
          faza.paint = 3
          otc = Otc.select().where(Otc.detail == faza,Otc.oper == 'paint').first()
          if otc == None:
            otc = Otc.create(detail=faza,start=datetime.today(),oper='paint')
          else:
            otc.fix = 0
            otc.error += 1
            otc.save()
          faza.save()
          otc = Worker.select(User.telegram).join(User).where(Worker.oper.in_(['otc'])).tuples()
          await AllReport(otc,f'{faza.detail} {detail.worker_1.user.surname} Наряд на проверку покраски')
          return 'Покраска завершина'
      else:
        user = Worker.select().join(User).where(User.telegram == str(message.chat.id),Worker.oper == detail.oper).first()
        if user == None:
          return 'У вас нет такой операции'
        else:
          detail.worker_2 = user
          detail.save()
          return f'Вы присоединились к наряду {text[1]}'
    # elif detail.faza
    else:
      user = Worker.select().join(User).where(User.telegram == message.chat.id,Worker.oper == detail.oper).first()
      if user == None:
        return f'У вас нет такой операции'
      if detail.oper in ['paint','weld']:
        if Otc.select().join(Faza).where(Otc.detail == detail.faza,Otc.end == None,Otc.fix == 0).first() != None:
          return 'Наряд не прошел ОТК'
      detail.start = date
      detail.worker_1 = user
      detail.save()
      faza = detail.faza
      norm = round(detail.norm,2)
      sec = 60 / 100 * (int(str(norm).split('.')[1]))
      hour = int(str(norm).split('.')[0]) // 60
      min = int(str(norm).split('.')[0]) % 60
      if min < 10:
        min = f'0{min}'
      if detail.oper == 'set':
        faza.set = 2
        faza.save()
        return f'Наряд {text[1]} принят в комплектовку'
      elif detail.oper == 'assembly':
        faza.assembly = 2
        faza.save()
        return f'Наряд {text[1]} принят в сборку! Норма {hour}:{min}:{round(sec)}'
      elif detail.oper == 'weld':
        faza.weld = 2
        faza.save()
        return f'Наряд {text[1]} принят на сварку! Норма {hour}:{min}:{round(sec)}'
      elif detail.oper == 'paint':
        faza.paint = 2
        faza.save()
        return f'Наряд {text[1]} принят в покраску'
  return 'QR-код не распознан'




async def UserAdd(message):
  text = message.text.replace('user ','').replace('User ','').replace(' ','')
  text = text.split('\n')
  qr_user = []
  for i in text:
    user = User.select().where(User.surname == (i.split('-')[0]).capitalize())
    if len(user) == 1:
      user = user.first()
      worker = Worker.select().where(Worker.user == user,Worker.oper == (i.split('-')[1]).capitalize()).first()
      if worker == None:
        oper = Worker.select().where(Worker.oper == (i.split('-')[1]).capitalize()).first()
        worker = Worker.create(user=user,oper=oper.oper,oper_rus=oper.oper_rus)
        qr_user.append(worker.id)
      else:
        qr_user.append(worker.id)
  pdf = QR_pdf(qr_user)
  await ViktorDocs(pdf)
  return 'Документ отправлен'



  # work = Worker.select().join(User).where(Worker.oper_rus == text[2],User.surname == text[1]).first()
  # if work != None:
  #   return 'Операция работнику уже назначена'
  # else:
  #   user = User.select().where(User.surname == text[1]).first()
  #   if user == None:
  #     return 'Работник не найден'
  #   else:
  #     oper = Worker.select().where(Worker.oper_rus == text[2]).first()
  #     if oper == None:
  #       if len(text) == 4:
  #         Worker.create(user=user,oper=text[3],oper_rus=text[2])
  #         return 'Операция создана'
  #       else:
  #         return 'Операция не найдена'
  #     else:
  #       Worker.create(user=user,oper=oper.oper,oper_rus=oper.oper_rus)
  #       return 'Операция создана'






async def OtcUser(message,text,worker):
  detail = Otc.select().join(Faza).where(Faza.detail == text[1],Otc.end == None).first()

  if detail and detail.fix == False:
    return detail

  elif detail and detail.fix == True:
    if message.chat.id == 2037378584 or message.chat.id == 262285696:
      paint = Detail.select().where(Detail.detail == text[1],Detail.oper == 'paint',Detail.start != None,Detail.end == None).first()
      if paint:
        faza = paint.faza
        paint.end = datetime.today()
        paint.save()
        detail.end = datetime.today()
        detail.worker = 131
        detail.fix = 0
        detail.error += 1
        detail.save()
        faza.paint = 3
        faza.packed = 1
        faza.save()
        return f'Покраска сдана. Наряд проверен!'
      else:
        return f'Наряд отправлен на доработку'



  else:
    if message.chat.id == 2037378584 or message.chat.id == 262285696:
    # if message.chat.id == 365803624:
      paint = Detail.select().where(Detail.detail == text[1],Detail.oper == 'paint',Detail.start != None,Detail.end == None).first()
      if paint:
        faza = paint.faza
        paint.end = datetime.today()
        paint.save()
        otc = Otc.create(detail=faza,start=datetime.today(),oper='paint')
        faza.paint = 3
        faza.save()
        return f'Покраска сдана'
    details = Otc.select().join(Faza).where(Faza.detail == text[1],Otc.end != None)
    if len(details) == 0:
      return f'Наряд {text[1]} не сдан на проверку'

    elif len(details) == 1:
      result = ''
      for detail in details:
        if detail.oper == 'weld':
          result += f'Проверка сварки проведена: {detail.worker.user.surname} {detail.worker.user.name} {detail.worker.user.patronymic}\nНа покраску не сдавалась'
        if detail.oper == 'paint':
          result += f'Проверка покраски проведена: {detail.worker.user.surname} {detail.worker.user.name} {detail.worker.user.patronymic}\nНа сварку не сдавалась'
      return result

    elif len(details) == 2:
      result = ''
      for detail in details:
        if detail.oper == 'weld':
          result += f'Проверка сварки проведена: {detail.worker.user.surname} {detail.worker.user.name} {detail.worker.user.patronymic}\n'
        if detail.oper == 'paint':
          result += f'Проверка покраски проведена: {detail.worker.user.surname} {detail.worker.user.name} {detail.worker.user.patronymic}\n'
      return result
      



    # if Otc.select().join(Faza).where(Faza.detail == text[1],Otc.oper == 'paint').first() != None:
    #   return f'Проверка покраски проведена: {detail.worker.user.surname} {detail.worker.user.name} {detail.worker.user.patronymic}'
    # elif Otc.select().join(Faza).where(Faza.detail == text[1],Otc.oper == 'weld').first() != None:
    #   return f'Проверка сварки проведена: {detail.worker.user.surname} {detail.worker.user.name} {detail.worker.user.patronymic}'
    # else:
    #   return f'Наряд {text[1]} не сдан на проверку'


  # elif detail and detail.end != None:
  #   return f'Проверка проведена: {detail.worker.user.surname} {detail.worker.user.name} {detail.worker.user.patronymic}'
  # else: 
  #   return f'Наряд {text[1]} не сдан на проверку'

  #   if Otc.select().join(Faza).where(Faza.detail == text[1],Otc.end == None).first()


  #   detail.end = datetime.today()
  #   detail.worker = worker
  #   detail.save()
  #   faza = detail.detail
  #   faza.packed = 1
  #   faza.save()
  #   return f'Проверка покраски наряда {detail.detail.detail} пройдена'
  # else:
  #   detail = Otc.select().join(Faza).where(Faza.detail == text[1],Otc.end != None,Otc.oper == 'paint').first()
  #   if detail:
  #     return f'Проверка уже пройдена! Проверяющий: {detail.worker.user.surname} {detail.worker.user.name}'
  #   else:
  #     return f'Наряд {text[1]} не здан на проверку'







def AdminUser(message,text,worker):
  detail = Detail.select().where(Detail.detail == text[1],Detail.to_work == 1,Detail.start != None,Detail.end == None,Detail.oper == 'paint').first()
  if detail.worker_1.user.id == 92:
    return 'Маляр Лоншакова'
  if detail:
    detail.end = datetime.today()
    detail.save()
    faza = detail.faza
    if detail.oper == 'set' and Detail.select().where(Detail.oper == 'assembly',Detail.detail == detail.detail).first() != None:
      Detail.update({Detail.to_work: 1}).where(Detail.oper == 'assembly',Detail.detail == detail.detail).execute()
      faza.set = 3
      faza.assembly = 1
      faza.save()
      return 'Комплектовка завершина'
    elif detail.oper == 'set' and Detail.select().where(Detail.oper == 'assembly',Detail.detail == detail.detail).first() == None:
      Detail.update({Detail.to_work: 1}).where(Detail.oper == 'paint',Detail.detail == detail.detail).execute()
      faza.set = 3
      faza.assembly = 3
      faza.weld = 3
      faza.paint = 1
      faza.save()
      return 'Комплектовка завершина'
    elif detail.oper == 'assembly':
      Detail.update({Detail.to_work: 1}).where(Detail.oper == 'weld',Detail.detail == detail.detail).execute()
      faza.assembly = 3
      faza.weld = 1
      faza.save()
      return 'Сборка завершина'
    elif detail.oper == 'weld':
      Detail.update({Detail.to_work: 1}).where(Detail.oper == 'paint',Detail.detail == detail.detail).execute()
      faza.weld = 3
      faza.paint = 1
      faza.save()
      return 'Сварка завершина'
    elif detail.oper == 'paint':
      faza.paint = 3
      # faza.packed = 1
      Otc.create(detail=faza,start=datetime.today(),oper='paint')
      faza.save()
      return 'Покраска завершина'
  else:
    return f'Не нашел наряд {text[1]}'

async def InfoDetail(text):
  str = text
  detail = PointPart.select(PointPart,fn.MAX(Part.length).alias('length')).join(Part).join_from(PointPart,Point).where(PointPart.detail == str[1]).group_by(PointPart.detail)
  text = ''
  if len(detail) == 1:
    d = detail.first()
    text += f'Наряд {d.detail}\n'
    text += f'Длина {d.length}\n'
    text += f'Вес {d.part.assembly.weight}\n'
    text += f'Заказ {d.part.assembly.cas.cas}\n'
    text += f'Фаза {d.point.faza}\n'
    text += f'Марка {d.part.assembly.assembly}\n'
    text += f'Чертёж {d.point.draw}\n'
  set = Detail.select().where(Detail.detail == str[1],Detail.oper == 'set').first()
  if set != None:
    text += '----------КОМПЛЕКТОВКА----------\n'
    if set.start == None:
      text += 'Не проходила\n'
    else:
      text += f'Начата {set.start}\n'
      if set.end == None:
        text += f'Не закончена\n'
      else:
        text += f'Закончена {set.end}\n'
      text += f'{set.worker_1.user.surname}\n'
      if set.worker_2 != None:
        text += f'{set.worker_2.user.surname}\n'
  set = Detail.select().where(Detail.detail == str[1],Detail.oper == 'assembly').first()
  if set != None:
    text += '-------------СБОРКА-------------\n'
    if set.start == None:
      text += 'Не проходила\n'
    else:
      text += f'Начата {set.start}\n'
      if set.end == None:
        text += f'Не закончена\n'
      else:
        text += f'Закончена {set.end}\n'
      text += f'{set.worker_1.user.surname}\n'
      if set.worker_2 != None:
        text += f'{set.worker_2.user.surname}\n'
  set = Detail.select().where(Detail.detail == str[1],Detail.oper == 'weld').first()
  if set != None:
    text += '-------------СВАРКА-------------\n'
    if set.start == None:
      text += 'Не проходила\n'
    else:
      text += f'Начата {set.start}\n'
      if set.end == None:
        text += f'Не закончена\n'
      else:
        text += f'Закончена {set.end}\n'
      text += f'{set.worker_1.user.surname}\n'
      if set.worker_2 != None:
        text += f'{set.worker_2.user.surname}\n'
  otc = Otc.select().join(Faza).where(Otc.oper == 'weld',Faza.detail == str[1]).first()
  if otc != None:
    text += '-------------ОТК СВАРКА-------------\n'
    if otc.end == None:
      if otc.fix == 1:
        text += 'Отправлен на доработку\n'
      else:
        text += 'Не принята\n'
    else:
      text += f'Сдана {otc.end}\n'
      text += f'Принял {otc.worker.user.surname}\n'
  set = Detail.select().where(Detail.detail == str[1],Detail.oper == 'paint').first()
  if set != None:
    text += '-------------ПОКРАСКА-------------\n'
    if set.start == None:
      text += 'Не проходила\n'
    else:
      text += f'Начата {set.start}\n'
      if set.end == None:
        text += f'Не закончена\n'
      else:
        text += f'Закончена {set.end}\n'
      text += f'{set.worker_1.user.surname}\n'
      if set.worker_2 != None:
        text += f'{set.worker_2.user.surname}\n'
  otc = Otc.select().join(Faza).where(Otc.oper == 'paint',Faza.detail == str[1]).first()
  if otc != None:
    text += '-------------ОТК ПОКРАСКА-------------\n'
    if otc.end == None:
      if otc.fix == 1:
        text += 'Отправлен на доработку\n'
      else:
        text += 'Не принята\n'
    else:
      text += f'Сдана {otc.end}\n'
      text += f'Принял {otc.worker.user.surname}\n'
  ship = DetailPack.select().join(Faza).where(Faza.detail == str[1]).first()
  if ship != None:
    text += '-------------УПАКОВКА-------------\n'
    text += f'Пачка {ship.pack.number}\n'
    text += f'Упакована {ship.pack.date}\n'
    if ship.pack.shipment != None:
      text += '-------------ОТПРАВКА-------------\n'
      text += f'Номер машины {ship.pack.shipment.number}\n'
      text += f'Отправлен {ship.pack.shipment.date}\n'
  return text




def Birk(file_name):
  PrintBirk.delete().execute()
  wb = load_workbook(file_name,data_only=True)
  sheet = wb.get_sheet_by_name('Лист1')
  post = []
  for y in range(1,302):
    post.append(sheet.cell(row=y,column=1).value)

  detail = PointPart.select(PointPart,fn.MAX(Part.length).alias('length')).join(Part).join_from(PointPart,Point).where(PointPart.detail.in_(post)).group_by(PointPart.detail)
  pr = []
  for d in detail:
    pr.append((d.detail,d.length,d.part.assembly.weight,d.part.assembly.cas.cas,d.point.faza,d.part.assembly.assembly,'Run '+str(d.detail)+' weld '+d.part.assembly.cas.cas,d.point.draw,2))
  PrintBirk.insert_many(pr, fields=[PrintBirk.detail,PrintBirk.length,PrintBirk.weight,PrintBirk.case,PrintBirk.faza,PrintBirk.mark,PrintBirk.qr,PrintBirk.draw,PrintBirk.count]).execute()

  z = PrintBirk.select().order_by(PrintBirk.count,PrintBirk.detail)

  book = Workbook()
  sheet = book.active
  y = 1
  sheet.cell(row=y,column=1).value = 'наряд'
  sheet.cell(row=y,column=2).value = 'длина'
  sheet.cell(row=y,column=3).value = 'вес'
  sheet.cell(row=y,column=4).value = 'заказ'
  sheet.cell(row=y,column=5).value = 'фаза'
  sheet.cell(row=y,column=6).value = 'марка'
  sheet.cell(row=y,column=7).value = 'qr'
  sheet.cell(row=y,column=8).value = 'чертёж'
  sheet.cell(row=y,column=9).value = 'Количество'
  for i in z:
    y += 1
    sheet.cell(row=y,column=1).value = i.detail
    sheet.cell(row=y,column=2).value = i.length
    sheet.cell(row=y,column=3).value = float(i.weight)
    sheet.cell(row=y,column=4).value = i.case
    sheet.cell(row=y,column=5).value = i.faza
    sheet.cell(row=y,column=6).value = i.mark
    sheet.cell(row=y,column=7).value = i.qr
    sheet.cell(row=y,column=8).value = i.draw
    sheet.cell(row=y,column=9).value = i.count
  book.save('Пример.xlsx')
  return 'Пример.xlsx'

def DetailUserAdd(task):
  data = []
  if task.worker_2 != None:
    data.append((task.id,task.worker_1,task.faza.weight / 2,task.norm / 2))
    data.append((task.id,task.worker_2,task.faza.weight / 2,task.norm / 2))
  else:
    data.append((task.id,task.worker_1,task.faza.weight,task.norm))
  print(data)
  with connection.atomic():
    DetailUser.insert_many(data,fields=[DetailUser.detail,DetailUser.worker,DetailUser.weight,DetailUser.norm]).execute()

def OpenDoor(message):
  import socket
  if message.text == '/open':
    msg = '4,1'
  elif message.text == '/close':
    msg = '5,1'
  server = socket.socket()
  server.connect(('192.168.0.115',2000))
  server.send(msg.encode('utf-8'))
  return