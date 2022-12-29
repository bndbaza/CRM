from openpyxl import load_workbook, Workbook
import glob
from models import *
from datetime import datetime
from peewee import fn, JOIN, Case
import csv


def Dlist(id):
  ord = id
  case = Order.get_or_create(cas=ord)[0]
  excel = {}
  dates = datetime.today()
  wb = load_workbook(f'{ord}.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('Диспетчерский')
  mark = ''
  draw = ''
  markst = ''
  count = 0
  paint = ''
  for i in range(6,300):
    if sheet.cell(row=i,column=3).value == None:
      continue
    if sheet.cell(row=i,column=1).value != None:
      mark = (sheet.cell(row=i,column=1).value).strip()
      draw = (sheet.cell(row=i,column=2).value)
      count = int((sheet.cell(row=i,column=24).value))
      paint = str(sheet.cell(row=i,column=23).value)
      drawing = Drawing.create(assembly = mark,weight = 0,area = 0,cas = case,create_date = dates,count = count,more = 0,paint = paint)
      for c in range(count):
        point = Point.create(assembly = drawing,point_x = 0,point_y = 0,point_z = 0,name = 'Сито',faza = 1,draw = draw,create_date = dates)
      excel[drawing] = []

    if sheet.cell(row=i,column=9).value != None:
      markst = sheet.cell(row=i,column=9).value

    if sheet.cell(row=i,column=22).value != None:
      excel[drawing].append(sheet.cell(row=i,column=22).value)


    profile = Size(sheet.cell(row=i,column=5).value)

    manipulation = ''
    if sheet.cell(row=i,column=15).value != None:
      manipulation += 'скос,'
    if sheet.cell(row=i,column=16).value != None:
      manipulation += 'вырез,'
    if sheet.cell(row=i,column=17).value != None:
      manipulation += 'фаск,'
    if sheet.cell(row=i,column=18).value != None:
      manipulation += 'фрез,'
    if sheet.cell(row=i,column=19).value != None:
      manipulation += 'ток,'
    if sheet.cell(row=i,column=20).value != None:
      manipulation += 'гиб,'
    
    part = Part.create(
        assembly=drawing,
        number=0,
        count=int(sheet.cell(row=i,column=4).value),
        profile=profile[0],
        size=profile[1],
        length=int(sheet.cell(row=i,column=6).value),
        weight=float(sheet.cell(row=i,column=7).value),
        mark=markst,
        manipulation=manipulation,
        work=profile[2],
        width = profile[3],
        perimeter=0,
        depth=0,
        create_date=dates,
        sn=sheet.cell(row=i,column=3).value
      )
    if sheet.cell(row=i,column=14).value != None:
      for hol in (sheet.cell(row=i,column=14).value).split(';'):
        h = hol.replace('/Ø','Ø').split('Ø')
        # Hole.create(part=part,float(diameter=h[1]),count=int(h[0])/part.count,depth=20)
        Hole.create(part=part,diameter=float(h[1]),count=int(h[0]),depth=20)

  welds = {}
  for i in excel:
    welds[i] = {}
    for y in excel[i]:
      y = y.replace(' ','').replace(')','').split('(')
      if welds[i].get(y[1]) == None:
        welds[i][y[1]] = round(float(y[0].replace(',','.')))
      else:
        welds[i][y[1]] += round(float(y[0].replace(',','.')))
  for weld in welds:
    for w in welds[weld]:
      Weld.create(assembly=weld,cathet=w,length=welds[weld][w],count=1)
  CalcDlist(ord)

def CalcDlist(ord):
  drawing = Drawing.select(Drawing.id).join(Order).where(Order.cas == ord).tuples()
  part = Part.select().join(Drawing).join(Order).where(Order.cas == ord).group_by(
    Part.profile,Part.size,Part.width,Part.length,Part.weight,Part.mark,Part.manipulation,Part.assembly
  )
  index = 1
  for i in part:
    Part.update({Part.number: index}).where(
      # Part.assembly.in_(drawing),
      Part.assembly == i.assembly,
      Part.profile == i.profile,
      Part.size == i.size,
      Part.width == i.width,
      Part.length == i.length,
      Part.weight == i.weight,
      Part.mark == i.mark,
      Part.manipulation == i.manipulation
    ).execute()
    index += 1
  part2 = Part.select(Part,fn.SUM(Part.weight * Part.count).alias('weight_dr')).where(Part.assembly.in_(drawing)).group_by(Part.assembly)
  for y in part2:
    y.assembly.weight = y.weight_dr
    y.assembly.save()



def Size(str):
  print(str)
  str = str.replace('x','х').strip()
  if str.startswith('Тр.Ø'):
    if float(str.replace('Тр.Ø','').split('х')[0]) >= 273:
      i = ('Труба круглая',str.replace('Тр.Ø',''),'saw_b','')
    else:
      i = ('Труба круглая',str.replace('Тр.Ø',''),'saw_s','')
    return (i)

  elif str.startswith('Тр.кв.'):
    dic = str.replace('Тр.кв.','').replace(' ','').split('х')
    if dic[0] == dic[1]:
      text = f'{dic[0]}х{dic[2]}'
    else:
      text = f'{dic[0]}х{dic[1]}х{dic[2]}'
    if float(dic[0]) >= 273:
      i = ('Труба профильная',text,'saw_b','')
    else:
      i = ('Труба профильная',text,'saw_s','')
    return (i)
  
  elif str.startswith('Тр.пр.'):
    dic = str.replace('Тр.пр.','').replace(' ','').split('х')
    if dic[0] == dic[1]:
      text = f'{dic[0]}х{dic[2]}'
    else:
      text = f'{dic[0]}х{dic[1]}х{dic[2]}'
    if float(dic[0]) >= 273:
      i = ('Труба профильная',text,'saw_b','')
    else:
      i = ('Труба профильная',text,'saw_s','')
    return (i)
  
  elif str.startswith('тр.кв.'):
    dic = str.replace('тр.кв.','').split('х')
    if dic[0] == dic[1]:
      text = f'{dic[0]}х{dic[2]}'
    else:
      text = f'{dic[0]}х{dic[1]}х{dic[2]}'
    if float(dic[0]) >= 273:
      i = ('Труба профильная',text,'saw_b','')
    else:
      i = ('Труба профильная',text,'saw_s','')
    return (i)
  
  elif str.startswith('Тр.'):
    if float(str.replace('Тр.','').split('х')[0]) >= 273:
      i = ('Труба круглая',str.replace('Тр.',''),'saw_b','')
    else:
      i = ('Труба круглая',str.replace('Тр.',''),'saw_s','')
    return (i)
  
  elif str.startswith('тр.'):
    if float(str.replace('тр.','').split('х')[0]) >= 273:
      i = ('Труба круглая',str.replace('тр.',''),'saw_b','')
    else:
      i = ('Труба круглая',str.replace('тр.',''),'saw_s','')
    return (i)
  
  elif str.startswith('Шв.6,5'):
    i = ('Швеллер',str.replace('Шв.','') + 'П','saw_s','')
    return (i)
  
  elif str.startswith('Шв.'):
    if int(str.replace('Шв.','').replace('П','').replace('У','').replace(',','.')) >= 20:
      i = ('Швеллер',str.replace('Шв.','').replace('П','') + 'П','saw_b','')
    else:
      i = ('Швеллер',str.replace('Шв.','').replace('П','') + 'П','saw_s','')
    return (i)
  
  elif str.startswith('шв.'):
    if int(str.replace('шв.','').replace('П','').replace('У','').replace(',','.')) >= 20:
      i = ('Швеллер',str.replace('шв.','') + 'П','saw_b','')
    else:
      i = ('Швеллер',str.replace('шв.','') + 'П','saw_s','')
    return (i)

  elif str.startswith('Кр.Ø'):
    i = ('Круг',str.replace('Кр.Ø',''),'saw_s','')
    return (i)

  elif str.startswith('Круг '):
    i = ('Круг',str.replace('Круг ',''),'saw_s','')
    return (i)
  
  elif str.startswith('Квадрат '):
    i = ('Квадрат',str.replace('Квадрат ',''),'saw_s','')
    return (i)
  
  elif str.startswith('Рельс '):
    i = ('Рельс',str.replace('Рельс ',''),'saw_b','')
    return (i)

  elif str.startswith('Арм.Ø'):
    str = str.split()[0]
    i = ('Арматура',str.replace('Арм.Ø',''),'saw_s','')
    return (i)
  
  elif str.startswith('Ø'):
    str = str.split()[0]
    i = ('Арматура',str.replace('Ø',''),'saw_s','')
    return (i)
  
  elif str.startswith('Шестигранник '):
    i = ('Шестигранник',str.replace('Шестигранник ',''),'saw_s','')
    return (i)

  elif str.startswith('Ду.Ø'):
    if float(str.replace('Ду.Ø','').split('х')[0]) >= 273:
      i = ('Труба круглая',str.replace('Ду.Ø',''),'saw_b','')
    else:
      i = ('Труба круглая',str.replace('Ду.Ø',''),'saw_s','')
    return (i)
  
  elif str.startswith('Ду.'):
    if float(str.replace('Ду.','').split('х')[0]) >= 273:
      i = ('Труба круглая',str.replace('Ду.',''),'saw_b','')
    else:
      i = ('Труба круглая',str.replace('Ду.',''),'saw_s','')
    return (i)

  elif str.startswith('Лист '):  
    a = str.replace('Лист ','').split('х')
    if int(a[0]) > int(a[1]):
      i = ('Лист',a[1],'cgm',a[0])
    else:
      i = ('Лист',a[0],'cgm',a[1])
    return (i)
  
  elif str.startswith('лист '):  
    a = str.replace('лист ','').split('х')
    if int(a[0]) > int(a[1]):
      i = ('Лист',a[1],'cgm',a[0])
    else:
      i = ('Лист',a[0],'cgm',a[1])
    return (i)

  elif str.startswith('Уг.'):
    if int(str.replace('Уг.','').split('х')[0]) >= 200:
      i = ('Уголок',str.replace('Уг.',''),'saw_b','')
    else:
      i = ('Уголок',str.replace('Уг.',''),'saw_s','')
    return (i)
  
  elif str.startswith('уг.'):
    if int(str.replace('уг.','').split('х')[0]) >= 200:
      i = ('Уголок',str.replace('уг.',''),'saw_b','')
    else:
      i = ('Уголок',str.replace('уг.',''),'saw_s','')
    return (i)


  elif str.startswith('Резина '):  
    a = str.replace('Резина ','').split('х')
    if int(a[0]) > int(a[1]):
      i = ('Резина',a[1],'no_metal',a[0])
    else:
      i = ('Резина',a[0],'no_metal',a[1])
    return (i)


  elif str.startswith('Паронит '):  
    a = str.replace('Паронит ','').split('х')
    if int(a[0]) > int(a[1]):
      i = ('Паронит',a[1],'no_metal',a[0])
    else:
      i = ('Паронит',a[0],'no_metal',a[1])
    return (i)

  elif str.startswith('Стекло '):  
    a = str.replace('Стекло ','').split('х')
    if int(a[0]) > int(a[1]):
      i = ('Стекло',a[1],'no_metal',a[0])
    else:
      i = ('Стекло',a[0],'no_metal',a[1])
    return (i)

