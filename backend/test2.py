from openpyxl import Workbook
from models import *
from peewee import fn, Case
from datetime import datetime
import csv

def Time():
  drawing = []
  point = []
  popa = []
  with open('2325.23.xls','r', encoding='windows-1251',newline='') as file:
    reader = csv.reader(file, delimiter='\t')
    for row in reader:
      if row[3].replace(' ','') == '1' and row[0].replace(' ','') == 'DRAWING' and row[1].replace(' ','').find('(?)') == -1:
        drawing.append(row[1].replace(' ',''))
      # if row[0].replace(' ','') == 'ASSEMBLY' and row[1].replace(' ','') in drawing:
      if row[0].replace(' ','') == 'ASSEMBLY' and row[1].replace(' ','') == 'МАр-1':
        mark = row[1].replace(' ','')
        x = (row[3].replace(' ','')).split('/')[1]
        y = (row[3].replace(' ','')).split('/')[0]
        z = float(row[2].replace(' ','').replace('+','').replace('-','')),
        name = row[5].replace(' ','')
        paint = (1,)


        if mark in ('МПл-5','МПл-39'):
          paint = (8,)

        elif name.lower() in ('ригельфахверка','настил','фахверк','монорельс','ограждение','упор','прогон'):
          paint = (1,2)

        elif y in ('8','7','7-8','8-7'):
          paint = (1,)

        elif mark in ('МПл-70'):
          paint = (1,2)

        elif name.lower() in ('лестница'):
          paint = (1,2)

        elif y in ('1','2','3','4','5','6') and name.lower() == 'балка':
          paint = (1,2,3,6)

        elif x in ('А','Б','В','Г') and name.lower() == 'балка':
          paint = (1,2,3,6)

        elif name == 'Колонна':
          paint = (1,2,3)

        elif name.lower() in ('связьвертикальная','связьверт.'):
          paint = (1,2,3,6)

        elif z[0] > 24.5:
          paint = (1,2,3,6)

        elif y in ('1-2','2-3','3-4','4-5','5-6','6-7') and x in ('А-Б','Б-В','В-Г') and name.lower() in ('балка','ригель'):
          paint = (1,2,4,7)

        # popa = PointPaint.select(Point.id).join(Point).join(Drawing).where(Drawing.cas == 15).tuples()
        pp = Point.select().join(Drawing).where(Point.id.not_in(popa),Drawing.assembly == mark,Drawing.cas == 15).first()
        popa.append(pp.id)
        print(mark)
        for i in paint:
          point.append((pp,i,'0'))

          # PointPaint.create(point=pp,coat=i,number='0')

        # point.append((
        #   row[1].replace(' ',''),
        #   row[2].replace(' ',''),
        #   (row[3].replace(' ','')).split('/')[0],
        #   (row[3].replace(' ','')).split('/')[1],
        #   row[5].replace(' ',''),
        #   paint
        # ))
  with connection.atomic():
    PointPaint.insert_many(point, fields=[PointPaint.point,PointPaint.coat,PointPaint.number]).execute()
  return
  TestTime.insert_many(point,fields=[TestTime.mark,TestTime.z,TestTime.y,TestTime.x,TestTime.name,TestTime.paint]).execute()

def Time2():
  return
  book = Workbook()
  sheet = book.active
  tt = TestTime.select()
  index = 0
  for t in tt:
    index += 1
    sheet.cell(row=index,column=1).value = t.mark
    sheet.cell(row=index,column=2).value = t.x
    sheet.cell(row=index,column=3).value = t.y
    sheet.cell(row=index,column=4).value = t.z
    sheet.cell(row=index,column=5).value = t.name
    sheet.cell(row=index,column=6).value = t.paint
  book.save(f'paint1.xlsx')

def Time3():
  return
  book = Workbook()
  sheet = book.active
  tt = TestTime.select().group_by(TestTime.mark,TestTime.paint)
  index = 0
  for t in tt:
    index += 1
    sheet.cell(row=index,column=1).value = t.mark
    sheet.cell(row=index,column=2).value = t.x
    sheet.cell(row=index,column=3).value = t.y
    sheet.cell(row=index,column=4).value = t.z
    sheet.cell(row=index,column=5).value = t.name
    sheet.cell(row=index,column=6).value = t.paint
  book.save(f'paint2.xlsx')

def NormPaint():
  details = Detail.select().join(Faza).where(Faza.case.in_((26,)),Detail.oper == 'paint')
  for detail in details:
    paint = 0
    pp = PointPart.select().where(PointPart.detail == detail.detail).group_by(PointPart.point)
    for p in pp:
      popa = PointPaint.select(fn.SUM(Coating.price * Drawing.area)).join(Point).join(Drawing).join_from(PointPaint,Coating).where(Point.id == p.point.id).scalar()
      paint += popa
    detail.norm = paint
    detail.save()
    print(detail.detail)


def UserNorm():
  start = datetime(2022,7,1)
  end = datetime(2022,8,1)
  users = Worker.select().where((Worker.oper == 'paint') | (Worker.oper == 'master'))
  for user in users:
    detail = Detail.select(fn.SUM(Case(None,[(Detail.worker_2 != None,Detail.norm / 2)],Detail.norm))).where((Detail.worker_1 == user) | (Detail.worker_2 == user),Detail.end != None,Detail.end > start,Detail.end < end ).scalar()
    detail = (str(detail).split('.')[0])
    print(user.user.surname,detail)

def Time4():
  lis = []
  # pp = Point.select().join(Drawing).where(Drawing.cas.in_((19,20,21,23,24,25,27,29)))
  # for p in pp:
  #   lis.append((p,1,'0'))
  #   lis.append((p,2,'0'))
  #   lis.append((p,2,'0'))
  # pp = Point.select().join(Drawing).where(Drawing.cas == 30)
  # for p in pp:
  #   lis.append((p,9,'0'))
  # pp = Point.select().join(Drawing).where(Drawing.cas == 28)
  # for p in pp:
  #   lis.append((p,10,'0'))
  ppaint = PointPaint.select(Point.id).join(Point).tuples()
  pp = Point.select().join(Drawing).where(Drawing.cas == 26,Point.id.not_in(ppaint))
  for p in pp:
    print(p)
    lis.append((p,8,'0'))
  
  with connection.atomic():
    PointPaint.insert_many(lis, fields=[PointPaint.point,PointPaint.coat,PointPaint.number]).execute()

  
  
  