from models import Drawing, Faza, Hole, Order, Point, Part, PointPart, Detail, AssemblyNorm, SawNorm, TaskPart, Weld, WeldNorm, HoleNorm, Task, TaskPart
from db import connection
from peewee import fn, JOIN
from openpyxl import load_workbook

def Faza_update(order):
  post = []
  points = Point.select().where(Order.cas == order,Point.line == None).join(Drawing).join(Order).order_by(Point.point_z,Point.point_y,Point.point_x)
  line1 = Point.select(fn.MAX(Point.line).alias('line'),fn.MAX(Point.faza).alias('faza')).where(Order.cas == order).join(Drawing).join(Order).order_by(Point.point_z,Point.point_y,Point.point_x).first()
  weight = Point.select(fn.SUM(Drawing.weight).alias('weight')).where(Order.cas == order,Point.faza == line1.faza).join(Drawing).join(Order).order_by(Point.point_z,Point.point_y,Point.point_x).first()
  if line1.line != None:
    line = int(line1.line) + 1 
    faza=line1.faza
  else:
    line = 1
    faza = 1
  weight=weight.weight
  for row in points:
    row.line = line
    line += 1
    weight += row.assembly.weight
    if weight > 15005:
      weight = row.assembly.weight
      faza += 1
    row.faza = faza
    post.append(row)
  Point.bulk_update(post, fields=[Point.line,Point.faza])



def Faza_update_test(order):
  wb = load_workbook('faza.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('2313.3')
  for i in range(1,145):
    try:
      drawing = Drawing.select().join(Order).where(Order.cas == order,Drawing.assembly == (sheet.cell(row=i,column=1).value)).first()
      point = Point.update({Point.faza: sheet.cell(row=i,column=3).value}).where(Point.assembly == drawing)
      return point[0]
    except:
      pass

def Faza_update_garage(order):
  drawing  = Drawing.select().join(Order).where(Order.cas == order)
  # Point.update({Point.faza: 1}).where(Point.assembly << drawing,Point.point_x.cast('int') <= 3).execute()
  # Point.update({Point.faza: 4}).where(Point.assembly == 559,Point.point_x.cast('int') <= 3).execute()
  # Point.update({Point.faza: 2}).where(Point.assembly == 577,Point.point_x.cast('int') < 10).execute()
  # Point.update({Point.faza: 2}).where(Point.assembly == 578,Point.point_x.cast('int') < 10).execute()
  # Point.update({Point.faza: 2}).where(Point.assembly == 579,Point.point_x.cast('int') < 10).execute()
  # Point.update({Point.faza: 2}).where(Point.assembly << drawing,Point.point_x.cast('int') > 3,Point.point_x.cast('int') <= 10).execute()
  # Point.update({Point.faza: 3}).where(Point.assembly == 577,Point.point_x.cast('int') >= 10).execute()
  # Point.update({Point.faza: 3}).where(Point.assembly == 578,Point.point_x.cast('int') >= 10).execute()
  # Point.update({Point.faza: 3}).where(Point.assembly == 579,Point.point_x.cast('int') >= 10).execute()
  # Point.update({Point.faza: 3}).where(Point.assembly << drawing,Point.point_x.cast('int') > 10).execute()
  # Point.update({Point.faza: 5}).where(Point.assembly << drawing,Point.faza == None).execute()
  Point.update({Point.faza: 4}).where(Point.assembly << drawing,Point.faza == 5).execute()
  return



def PartPoint(faza,cas):
  parts = Part.filter(assembly__cas = cas)
  points = Point.filter(faza = faza,assembly__cas = cas)
  maxpp = PointPart.select(fn.MAX(PointPart.detail)).scalar()
  all = []
  if maxpp == None:
    maxpp = 1
  else:
    maxpp += 1
  max2 = maxpp
  for point in points:
    for part in parts:
      if point.assembly == part.assembly:
        oper = {'hole': 0, 'chamfer': 0, 'cgm': 0, 'saw': 0, 'milling': 0, 'notch': 0, 'bevel': 0, 'joint':0,'bend':0,'turning':0}
        if part.chamfers.count() > 0:
          oper['chamfer'] = 1
        if part.holes.count() > 0:
          oper['hole'] = 1
        if part.profile == 'Лист':
          oper['cgm'] = 1
        elif part.profile == 'Паронит' or part.profile == 'Резина' or part.profile == 'Стекло':
          pass
        else:
          oper['saw'] = 1
        if part.manipulation.find('фрез') >= 0:
          oper['milling'] = 1
        if part.manipulation.find('вырез') >= 0:
          oper['notch'] = 1
        if part.manipulation.find('скос') >= 0:
          oper['bevel'] = 1
        if part.manipulation.find('стык') >= 0:
          oper['joint'] = 1
        if part.manipulation.find('гиб') >= 0:
          oper['bend'] = 1
        if part.manipulation.find('ток') >= 0:
          oper['turning'] = 1
        if part.manipulation.find('резьба') >= 0:
          oper['turning'] = 1
        d = (point,part,maxpp,oper['hole'],oper['chamfer'],oper['cgm'],oper['saw'],oper['milling'],oper['notch'],oper['bevel'],oper['joint'],oper['bend'],oper['turning'])
        all.append(d)
    maxpp += 1
  with connection.atomic():
    PointPart.insert_many(all, fields=[PointPart.point,PointPart.part,PointPart.detail,PointPart.hole,PointPart.chamfer,PointPart.cgm,PointPart.saw,PointPart.milling,PointPart.notch,PointPart.bevel,PointPart.joint,PointPart.bend,PointPart.turning]).execute()
  Test(max2,faza,cas)

# def Test2():
#   z = PointPart.select(PointPart).join(Part).join(Drawing)
#   y = 1
#   case = z[0].point.assembly
#   all = []
#   for i in z:
#     if i.point.assembly != case:
#       print(i.point.assembly,case)
#       y += 1
#     i.detail = y
#     case = i.point.assembly
#     all.append(i)
#   PointPart.bulk_update(all, fields=[PointPart.detail])
  
def Test(max2,faza,case):
  # zi = PointPart.select(PointPart,fn.COUNT(PointPart.detail).alias('aaa')).group_by(PointPart.detail).having(fn.COUNT(PointPart.detail) != 1)
  # yi = PointPart.select(PointPart,fn.COUNT(PointPart.detail).alias('aaa')).group_by(PointPart.detail).having(fn.COUNT(PointPart.detail) == 1)
  # xi = PointPart.select()
  zi = PointPart.select(PointPart,fn.COUNT(PointPart.detail).alias('aaa')).join(Point).join(Drawing).where(Point.faza == faza,Drawing.cas == case).group_by(PointPart.detail).having(fn.COUNT(PointPart.detail) != 1)
  yi = PointPart.select(PointPart,fn.COUNT(PointPart.detail).alias('aaa')).join(Point).join(Drawing).where(Point.faza == faza,Drawing.cas == case).group_by(PointPart.detail).having(fn.COUNT(PointPart.detail) == 1)
  xi = PointPart.select().join(Point).join(Drawing).where(Point.faza == faza,Drawing.cas == case)
  # index = PointPart.select(fn.MAX(PointPart.detail)).scalar()
  # print(index)
  index = max2
  all = []
  for i in zi:
    i.detail = index
    index += 1

  dec = {}
  for i in yi:
    if i.part.profile+i.part.size in dec.keys():
      i.detail = dec[i.part.profile+i.part.size]
    else:
      dec[i.part.profile+i.part.size] = index
      i.detail = index
      index += 1

  for i in xi:
    for y in yi:
      if i.point == y.point:
        i.detail = y.detail
        i.weld = 0
        all.append(i)
        
  for i in xi:
    for y in zi:
      if i.point == y.point:
        i.detail = y.detail
        all.append(i)
  PointPart.bulk_update(all, fields=[PointPart.detail,PointPart.weld])

def Detail_create(faza,case):
  d = []
  # pointpart_work = PointPart.select(PointPart,
  #                                   fn.SUM(PointPart.hole),
  #                                   fn.SUM(PointPart.bevel),
  #                                   fn.SUM(PointPart.notch),
  #                                   fn.SUM(PointPart.chamfer),
  #                                   fn.SUM(PointPart.milling),
  #                                   fn.SUM(PointPart.bend)).join(Part).join(Drawing).join(Order).join_from(PointPart,Point).where(Order.cas == case,Point.faza == faza).group_by(PointPart.detail,Part.work)
  # for i in pointpart_work:
  #   if i.hole != 0:
  #     if i.part.work == 'saw_s' or i.part.work == 'saw_b':
  #       norm_h = 0
  #       for part in PointPart.select().where(PointPart.detail == i.detail):
  #         for hole in Hole.select().where(Hole.part == part.part.id):
  #           norm_h += (HoleNorm.select().where(HoleNorm.diameter >= hole.diameter,
  #                                             HoleNorm.lenght_of <= int(i.part.length),
  #                                             HoleNorm.lenght_to >= int(i.part.length),
  #                                             HoleNorm.depth_of <= i.part.depth,
  #                                             HoleNorm.depth_to >= i.part.depth,
  #                                             HoleNorm.count >= hole.count,
  #                                             HoleNorm.metal == 'Сорт').first()).norm * hole.count * i.part.count
  #       d.append((i.detail,i.part.work,'hole',norm_h))
  #     else:
  #       norm_h = 0
  #       for part in PointPart.select().where(PointPart.detail == i.detail):
  #         for hole in Hole.select().where(Hole.part == part.part.id):
  #           norm_h += (HoleNorm.select().where(HoleNorm.diameter >= hole.diameter,
  #                                           HoleNorm.depth_of <= i.part.depth,
  #                                           HoleNorm.depth_to >= i.part.depth,
  #                                           HoleNorm.count >= hole.count,
  #                                           HoleNorm.metal == 'Лист').first()).norm * hole.count * i.part.count                            
  #       d.append((i.detail,i.part.work,'hole',norm_h))
  #   if i.bevel != 0:
  #     d.append((i.detail,i.part.work,'bevel'))
  #   if i.notch != 0:
  #     d.append((i.detail,i.part.work,'notch'))
  #   if i.chamfer != 0:
  #     d.append((i.detail,i.part.work,'chamfer'))
  #   if i.milling != 0:
  #     d.append((i.detail,i.part.work,'milling'))
  #   if i.bend != 0:
  #     d.append((i.detail,i.part.work,'bend'))
  #   if i.part.work == 'cgm':
  #     d.append((i.detail,i.part.work,'cgm'))
  #   if i.part.work == 'saw_s':
  #     norm_s = 0
  #     for ss in PointPart.select(PointPart,Part).join(Part).where(PointPart.detail == i.detail,Part.work == 'saw_s'):
  #       norm_s += (SawNorm.select().where(SawNorm.profile == ss.part.profile,SawNorm.size == ss.part.size).get()).norm_direct
  #     d.append((i.detail,i.part.work,'saw',norm_s))
  #   if i.part.work == 'saw_b':
  #     norm_b = 0
  #     for ss in PointPart.select(PointPart,Part).join(Part).where(PointPart.detail == i.detail,Part.work == 'saw_b'):
  #       norm_b += (SawNorm.select().where(SawNorm.profile == ss.part.profile,SawNorm.size == ss.part.size).get()).norm_direct
  #     d.append((i.detail,i.part.work,'saw',norm_b))

    
  # pointpart_paint = PointPart.select().join(Point).join(Drawing).join(Order).where(Order.cas == case,Point.faza == faza).group_by(PointPart.detail)
  # data = []
  # for i in pointpart_paint:
  #   data.append((i.detail,faza,i.point.assembly.cas))
  # with connection.atomic():
  #   Faza.insert_many(data, fields=[Faza.detail,Faza.faza,Faza.case]).execute()

  data = []
  pointpart_paint = PointPart.select(PointPart,fn.SUM(Drawing.weight).alias('weight'),fn.SUM(PointPart.weld).alias('sum_weld')).join(Point).join(Drawing).where(Drawing.cas == case,Point.faza == faza).group_by(PointPart.detail)
  for i in pointpart_paint:
    if i.sum_weld == 0:
      x = 1
    else:
      x = i.sum_weld
    data.append((i.detail,i.point.faza,i.point.assembly.cas,i.weight / x))
  with connection.atomic():
    Faza.insert_many(data, fields=[Faza.detail,Faza.faza,Faza.case,Faza.weight]).execute()


  pointpart_weld = PointPart.select(PointPart,fn.SUM(Part.count).alias('count')).join(Point).join(Drawing).join_from(PointPart,Part).where(PointPart.weld == 1,Drawing.cas == case,Point.faza == faza).group_by(PointPart.detail)
  for i in pointpart_weld:
    faza_tab = Faza.get(Faza.detail == i.detail)
    
    norm_assembly = AssemblyNorm.select().where(AssemblyNorm.name == i.point.name,
                                                i.point.assembly.weight >= AssemblyNorm.mass_of,
                                                i.point.assembly.weight < AssemblyNorm.mass_to,
                                                i.count >= AssemblyNorm.count_of,
                                                i.count < AssemblyNorm.count_to).first()
    try:
      d.append((i.detail,'weld','assembly',norm_assembly.norm * (i.point.assembly.weight / 1000),faza_tab))
    except:
      d.append((i.detail,'weld','assembly',0,faza_tab))
      print(i.point.name)

    weld = Weld.select().where(Weld.assembly == i.point.assembly)
    norm_weld = 0
    for w in weld:
      weld = WeldNorm.select((WeldNorm.norm * w.length / 1000 / i.point.assembly.count).alias('aaa')).where(WeldNorm.cathet == w.cathet).first()
      norm_weld += weld.aaa
    d.append((i.detail,'weld','weld',norm_weld,faza_tab))

  for i in pointpart_paint:
    faza_tab = Faza.get(Faza.detail == i.detail)
    d.append((i.detail,'paint','paint',0,faza_tab))
    d.append((i.detail,'set','set',0,faza_tab))
  with connection.atomic():
    Detail.insert_many(d, fields=[Detail.detail,Detail.basic,Detail.oper,Detail.norm,Detail.faza]).execute()

def Task_create(faza,case):
  index = Task.select(fn.MAX(Task.task)).where(Task.faza < faza).scalar()
  if index == None: index = 0
  d = []
  tasks = PointPart.select(PointPart,Point,Part).join(Part).join(Drawing).join(Order).join_from(PointPart,Point).where(Order.id == case.id,Point.faza == faza).group_by(Part.profile,Part.size,Part.mark)
  for i in tasks:
    index += 1
    task = Task(task = index,oper = i.part.work,faza = faza, order = case)
    task.save()
    taskparts = PointPart.select(PointPart,Part,fn.SUM(Part.count).alias('count_task')).join(Part).join(Drawing).join(Order).join_from(PointPart,Point).where(Order.id == case.id,Point.faza == faza,Part.profile == i.part.profile,Part.size == i.part.size,Part.mark == i.part.mark).group_by(Part)
    for taskpart in taskparts:
      d.append((task,taskpart.part,taskpart.count_task))
  aaa = ((PointPart.hole,'hole'),(PointPart.bevel,'bevel'),(PointPart.notch,'notch'),(PointPart.chamfer,'chamfer'),(PointPart.milling,'milling'),(PointPart.bend,'bend'),(PointPart.turning,'turning'),(PointPart.joint,'joint'))
  for a in aaa:
    tasks_hole = PointPart.select(PointPart,Part,fn.SUM(Part.count).alias('count_task')).join(Part).join(Drawing).join(Order).join_from(PointPart,Point).where(Order.id == case.id,Point.faza == faza,a[0] == 1).group_by(Part.number)
    for hole in tasks_hole:
      index += 1
      task_hole = Task(task = index,oper = a[1],faza = faza,order = case)
      task_hole.save()
      tasks_part = PointPart.select(PointPart,Part,fn.SUM(Part.count).alias('count_task')).join(Part).join(Drawing).join(Order).join_from(PointPart,Point).where(Order.id == case.id,Point.faza == faza,a[0] == 1,Part.number == hole.part.number).group_by(Part)
      for part in tasks_part:
        d.append((task_hole,part.part,part.count_task))
  with connection.atomic():
    TaskPart.insert_many(d, fields=[TaskPart.task,TaskPart.part,TaskPart.count]).execute()