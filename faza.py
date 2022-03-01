from models import Drawing, Hole, Order, Point, Part, PointPart, Detail, AssemblyNorm, SawNorm, Weld, WeldNorm, HoleNorm
from db import connection
from peewee import fn, JOIN
from openpyxl import load_workbook

def Faza_update(order):
  post = []
  points = Point.select().where(Order.cas == order,Point.line == None).join(Drawing).join(Order).order_by(Point.point_z,Point.point_y,Point.point_x)
  line1 = Point.select(fn.MAX(Point.line).alias('line'),fn.MAX(Point.faza).alias('faza')).where(Order.cas == order).join(Drawing).join(Order).order_by(Point.point_z,Point.point_y,Point.point_x).first()
  weight = Point.select(fn.SUM(Drawing.weight).alias('weight')).where(Order.cas == order,Point.faza == line1.faza).join(Drawing).join(Order).order_by(Point.point_z,Point.point_y,Point.point_x).first()
  line = int(line1.line) + 1 
  faza=line1.faza
  weight=weight.weight
  for row in points:
    row.line = line
    line += 1
    weight += row.assembly.weight
    if weight > 15005:
      weight = row.assembly.weight
      faza += 1
    row.faza = faza
    post.append(row)
  Point.bulk_update(post, fields=[Point.line,Point.faza])



def Faza_update_test(order):
  wb = load_workbook('faza.xlsx',data_only=True)
  sheet = wb.get_sheet_by_name('2313.3')
  for i in range(1,145):
    try:
      drawing = Drawing.select().join(Order).where(Order.cas == order,Drawing.assembly == (sheet.cell(row=i,column=1).value)).first()
      point = Point.update({Point.faza: sheet.cell(row=i,column=3).value}).where(Point.assembly == drawing)
      return point[0]
    except:
      pass

def Faza_update_garage(order):
  drawing  = Drawing.select().join(Order).where(Order.cas == order)
  # Point.update({Point.faza: 1}).where(Point.assembly << drawing,Point.point_x.cast('int') <= 3).execute()
  # Point.update({Point.faza: 4}).where(Point.assembly == 559,Point.point_x.cast('int') <= 3).execute()
  # Point.update({Point.faza: 2}).where(Point.assembly == 577,Point.point_x.cast('int') < 10).execute()
  # Point.update({Point.faza: 2}).where(Point.assembly == 578,Point.point_x.cast('int') < 10).execute()
  # Point.update({Point.faza: 2}).where(Point.assembly == 579,Point.point_x.cast('int') < 10).execute()
  # Point.update({Point.faza: 2}).where(Point.assembly << drawing,Point.point_x.cast('int') > 3,Point.point_x.cast('int') <= 10).execute()
  # Point.update({Point.faza: 3}).where(Point.assembly == 577,Point.point_x.cast('int') >= 10).execute()
  # Point.update({Point.faza: 3}).where(Point.assembly == 578,Point.point_x.cast('int') >= 10).execute()
  # Point.update({Point.faza: 3}).where(Point.assembly == 579,Point.point_x.cast('int') >= 10).execute()
  # Point.update({Point.faza: 3}).where(Point.assembly << drawing,Point.point_x.cast('int') > 10).execute()
  # Point.update({Point.faza: 5}).where(Point.assembly << drawing,Point.faza == None).execute()
  Point.update({Point.faza: 4}).where(Point.assembly << drawing,Point.faza == 5).execute()
  return



def PartPoint(faza,cas):
  parts = Part.filter(assembly__cas__cas = cas)
  points = Point.filter(faza = faza,assembly__cas__cas = cas)
  maxpp = PointPart.select(fn.MAX(PointPart.detail)).scalar()
  all = []
  if maxpp == None:
    maxpp = 1
  else:
    maxpp += 1
  max2 = maxpp
  for point in points:
    for part in parts:
      if point.assembly == part.assembly:
        oper = {'hole': 0, 'chamfer': 0, 'cgm': 0, 'saw': 0, 'milling': 0, 'notch': 0, 'bevel': 0}
        if part.chamfers.count() > 0:
          oper['chamfer'] = 1
        if part.holes.count() > 0:
          oper['hole'] = 1
        if part.profile == 'Лист':
          oper['cgm'] = 1
        else:
          oper['saw'] = 1
        if part.manipulation.find('фрез') >= 0:
          oper['milling'] = 1
        if part.manipulation.find('вырез') >= 0:
          oper['notch'] = 1
        if part.manipulation.find('скос') >= 0:
          oper['bevel'] = 1
        d = (point,part,maxpp,oper['hole'],oper['chamfer'],oper['cgm'],oper['saw'],oper['milling'],oper['notch'],oper['bevel'])
        all.append(d)
    maxpp += 1
  with connection.atomic():
    PointPart.insert_many(all, fields=[PointPart.point,PointPart.part,PointPart.detail,PointPart.hole,PointPart.chamfer,PointPart.cgm,PointPart.saw,PointPart.milling,PointPart.notch,PointPart.bevel]).execute()
  Test(max2,faza,cas)

# def Test2():
#   z = PointPart.select(PointPart).join(Part).join(Drawing)
#   y = 1
#   case = z[0].point.assembly
#   all = []
#   for i in z:
#     if i.point.assembly != case:
#       print(i.point.assembly,case)
#       y += 1
#     i.detail = y
#     case = i.point.assembly
#     all.append(i)
#   PointPart.bulk_update(all, fields=[PointPart.detail])
  
def Test(max2,faza,case):
  # zi = PointPart.select(PointPart,fn.COUNT(PointPart.detail).alias('aaa')).group_by(PointPart.detail).having(fn.COUNT(PointPart.detail) != 1)
  # yi = PointPart.select(PointPart,fn.COUNT(PointPart.detail).alias('aaa')).group_by(PointPart.detail).having(fn.COUNT(PointPart.detail) == 1)
  # xi = PointPart.select()
  zi = PointPart.select(PointPart,fn.COUNT(PointPart.detail).alias('aaa')).join(Point).join(Drawing).join(Order).where(Point.faza == faza,Order.cas == case).group_by(PointPart.detail).having(fn.COUNT(PointPart.detail) != 1)
  yi = PointPart.select(PointPart,fn.COUNT(PointPart.detail).alias('aaa')).join(Point).join(Drawing).join(Order).where(Point.faza == faza,Order.cas == case).group_by(PointPart.detail).having(fn.COUNT(PointPart.detail) == 1)
  xi = PointPart.select().join(Point).join(Drawing).join(Order).where(Point.faza == faza,Order.cas == case)
  # index = PointPart.select(fn.MAX(PointPart.detail)).scalar()
  # print(index)
  index = max2
  all = []
  for i in zi:
    i.detail = index
    index += 1
  dec = {}
  for i in yi:
    if i.part.profile+i.part.size in dec.keys():
      i.detail = dec[i.part.profile+i.part.size]
    else:
      dec[i.part.profile+i.part.size] = index
      i.detail = index
      index += 1
  for i in xi:
    for y in yi:
      if i.point == y.point:
        i.detail = y.detail
        i.weld = 0
        all.append(i)
  for i in xi:
    for y in zi:
      if i.point == y.point:
        i.detail = y.detail
        all.append(i)
  PointPart.bulk_update(all, fields=[PointPart.detail,PointPart.weld])

def Detail_create(faza,case):
  d = []
  pointpart_work = PointPart.select(PointPart,
                                    fn.SUM(PointPart.hole),
                                    fn.SUM(PointPart.bevel),
                                    fn.SUM(PointPart.notch),
                                    fn.SUM(PointPart.chamfer),
                                    fn.SUM(PointPart.milling),
                                    fn.SUM(PointPart.bend)).join(Part).join(Drawing).join(Order).join_from(PointPart,Point).where(Order.cas == case,Point.faza == faza).group_by(PointPart.detail,Part.work)
  for i in pointpart_work:
    if i.hole != 0:
      if i.part.work == 'saw_s' or i.part.work == 'saw_b':
        norm_h = 0
        for part in PointPart.select().where(PointPart.detail == i.detail):
          for hole in Hole.select().where(Hole.part == part.part.id):
            norm_h += (HoleNorm.select().where(HoleNorm.diameter >= hole.diameter,
                                              HoleNorm.lenght_of <= int(i.part.length),
                                              HoleNorm.lenght_to >= int(i.part.length),
                                              HoleNorm.depth_of <= i.part.depth,
                                              HoleNorm.depth_to >= i.part.depth,
                                              HoleNorm.count >= hole.count,
                                              HoleNorm.metal == 'Сорт').first()).norm * hole.count * i.part.count
        d.append((i.detail,i.part.work,'hole',norm_h))
      else:
        norm_h = 0
        for part in PointPart.select().where(PointPart.detail == i.detail):
          for hole in Hole.select().where(Hole.part == part.part.id):
            norm_h += (HoleNorm.select().where(HoleNorm.diameter >= hole.diameter,
                                            HoleNorm.depth_of <= i.part.depth,
                                            HoleNorm.depth_to >= i.part.depth,
                                            HoleNorm.count >= hole.count,
                                            HoleNorm.metal == 'Лист').first()).norm * hole.count * i.part.count                            
        d.append((i.detail,i.part.work,'hole',norm_h))
    if i.bevel != 0:
      d.append((i.detail,i.part.work,'bevel'))
    if i.notch != 0:
      d.append((i.detail,i.part.work,'notch'))
    if i.chamfer != 0:
      d.append((i.detail,i.part.work,'chamfer'))
    if i.milling != 0:
      d.append((i.detail,i.part.work,'milling'))
    if i.bend != 0:
      d.append((i.detail,i.part.work,'bend'))
    if i.part.work == 'cgm':
      d.append((i.detail,i.part.work,'cgm'))
    if i.part.work == 'saw_s':
      norm_s = 0
      for ss in PointPart.select(PointPart,Part).join(Part).where(PointPart.detail == i.detail,Part.work == 'saw_s'):
        norm_s += (SawNorm.select().where(SawNorm.profile == ss.part.profile,SawNorm.size == ss.part.size).get()).norm_direct
      d.append((i.detail,i.part.work,'saw',norm_s))
    if i.part.work == 'saw_b':
      norm_b = 0
      for ss in PointPart.select(PointPart,Part).join(Part).where(PointPart.detail == i.detail,Part.work == 'saw_b'):
        norm_b += (SawNorm.select().where(SawNorm.profile == ss.part.profile,SawNorm.size == ss.part.size).get()).norm_direct
      d.append((i.detail,i.part.work,'saw',norm_b))

  pointpart_weld = PointPart.select(PointPart,fn.SUM(Part.count).alias('count')).join(Point).join(Drawing).join(Order).join_from(PointPart,Part).where(PointPart.weld == 1,Order.cas == case,Point.faza == faza).group_by(PointPart.detail)
  for i in pointpart_weld:
    norm_assembly = AssemblyNorm.select().where(AssemblyNorm.name == i.point.name,
                                                i.point.assembly.weight >= AssemblyNorm.mass_of,
                                                i.point.assembly.weight < AssemblyNorm.mass_to,
                                                i.count >= AssemblyNorm.count_of,
                                                i.count < AssemblyNorm.count_to).first()
    try:
      d.append((i.detail,'weld','assembly',norm_assembly.norm * (i.point.assembly.weight / 1000)))
    except:
      print(i.point.name)
    weld = Weld.select().where(Weld.assembly == i.point.assembly)
    norm_weld = 0
    for w in weld:
      weld = WeldNorm.select((WeldNorm.norm * w.length / 1000 / i.point.assembly.count).alias('aaa')).where(WeldNorm.cathet == w.cathet).first()
      norm_weld += weld.aaa
    d.append((i.detail,'weld','weld',norm_weld))

  pointpart_paint = PointPart.select().join(Point).join(Drawing).join(Order).where(Order.cas == case,Point.faza == faza).group_by(PointPart.detail)
  for i in pointpart_paint:
    d.append((i.detail,'paint','paint'))
  with connection.atomic():
    Detail.insert_many(d, fields=[Detail.detail,Detail.basic,Detail.oper,Detail.norm]).execute()